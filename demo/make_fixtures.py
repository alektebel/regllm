"""Generate the demo Excel fixtures (and refresh the bundled UI assets).

Run from the repo root:  python demo/make_fixtures.py
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
FIXTURES = HERE / "fixtures"
ASSETS = HERE.parent / "DQC" / "app" / "src" / "assets" / "demo"


def diccionario() -> Path:
    wb = openpyxl.Workbook()
    notas = wb.active
    notas.title = "Notas"
    notas.append(["Diccionario de campos — ciclos de recuperación"])
    notas.append(["Versión", "2.3", "junio 2026"])
    dic = wb.create_sheet("DICCIONARIO")
    dic.append(["Field", "Type", "Description", "Null", "Formula"])
    for row in [
        ["PD_ESTIMADA", "REAL", "Probabilidad de default estimada (0-1)", "no", ""],
        ["LGD_ESTIMADA", "REAL", "Pérdida en caso de default estimada (0-1)", "no", ""],
        ["EAD_TOTAL", "REAL", "Exposición total en el momento del default (EUR)", "no",
         "EAD_BALANCE + EAD_FUERA_BALANCE"],
        ["ECL", "REAL", "Pérdida crediticia esperada", "no",
         "PD_ESTIMADA * LGD_ESTIMADA * EAD_TOTAL"],
        ["STAGE_IFRS9", "INTEGER", "Stage IFRS9 (1, 2 o 3)", "no", ""],
        ["FECHA_CIERRE", "DATE", "Fecha de cierre del ciclo", "si", ""],
    ]:
        dic.append(row)
    path = FIXTURES / "diccionario_demo.xlsx"
    wb.save(path)
    return path


def casos() -> Path:
    """Extracted cases with an ECL column (so the reperformance DQC can run)
    and DQC_ID labels for precision/recall. Violations by design:
      caso 1: PD > 1                       (label DQC_PD_001)
      caso 2: EAD < 0                      (label DQC_EAD_002)
      caso 3: ECL != PD*LGD*EAD            (label DQC_ECL_003)
      caso 4: PD > 1 y EAD < 0             (labels DQC_PD_001;DQC_EAD_002)
      caso 5: limpio
      caso 6: LGD > 1                      (sin label — la regla LGD no lleva id)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CASOS"
    ws.append(["PD_ESTIMADA", "EAD_TOTAL", "LGD_ESTIMADA", "ECL",
               "STAGE_IFRS9", "DQC_ID"])
    for row in [
        [1.15, 12000, 0.45, 6210.0, 2, "DQC_PD_001"],
        [0.02, -500, 0.40, -4.0, 1, "DQC_EAD_002"],
        [0.85, 30000, 0.55, 99999.0, 3, "DQC_ECL_003"],
        [1.30, -20, 0.35, -9.1, 2, "DQC_PD_001;DQC_EAD_002"],
        [0.10, 8000, 0.42, 336.0, 1, ""],
        [0.50, 1000, 1.20, 600.0, 2, ""],
    ]:
        ws.append(row)
    path = FIXTURES / "casos_demo.xlsx"
    wb.save(path)
    return path


def diccionario_ambiguo() -> Path:
    """TWO dictionary-shaped sheets with the same headers — triggers the
    ambiguous-sheet branch: the UI must ask which one to use."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, extra in [("CAMPOS_2024", "cartera 2024"),
                        ("CAMPOS_2025", "cartera 2025")]:
        ws = wb.create_sheet(name)
        ws.append(["Field", "Type", "Description"])
        ws.append(["PD_ESTIMADA", "REAL", f"PD estimada ({extra})"])
        ws.append(["EAD_TOTAL", "REAL", f"EAD total ({extra})"])
        ws.append(["LGD_ESTIMADA", "REAL", f"LGD estimada ({extra})"])
        ws.append(["ECL", "REAL", f"ECL ({extra})"])
        ws.append(["STAGE_IFRS9", "INTEGER", f"Stage ({extra})"])
    path = FIXTURES / "diccionario_ambiguo.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    FIXTURES.mkdir(exist_ok=True)
    paths = [diccionario(), casos(), diccionario_ambiguo()]
    ASSETS.mkdir(parents=True, exist_ok=True)
    for p in paths[:2]:                      # the two the UI auto-loads
        shutil.copy(p, ASSETS / p.name)
    for p in paths:
        print("wrote", p)
    print("assets refreshed in", ASSETS)
