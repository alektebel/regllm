#!/usr/bin/env python3
"""Analyze any CSV, build embedding weights, and generate SAS include code.

Auto-detects column roles (categorical / numeric / skip / metadata) and
produces everything needed for the SAS embedding pipeline:

  1. Projection matrix (59→50 by default, CTX×EMBED_DIM in general)
  2. Categorical encoding maps → CSV
  3. Numeric normalization params → CSV
  4. sas_generated.sas → DATA step + IML code specific to this table

Usage:
    # Auto-detect columns
    python scripts/export_embedding_weights.py --csv my_data.csv

    # Manual column roles  
    python scripts/export_embedding_weights.py --csv my_data.csv \\
        --cat-cols SEGMENTO PRODUCTO --num-cols PD_ESTIMADA EAD \\
        --meta-cols ID_CONTR --skip-cols OBSOLETE_FLAG
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "embeddings" / "tabular"


def _load_csv(path: Path) -> tuple[list[str], list[dict]]:
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        rows = list(reader)
    return cols, rows


def _is_numeric(col: str, rows: list[dict], sample: int = 1000) -> bool:
    """Check if a column is numeric by trying to parse a sample of values."""
    successes = 0
    trials = 0
    for r in rows[:sample]:
        v = r.get(col, "").strip()
        if v == "":
            continue
        trials += 1
        try:
            float(v)
            successes += 1
        except ValueError:
            pass
    if trials == 0:
        return False
    return successes / trials >= 0.8


def _is_id_like(col: str) -> bool:
    """Detect identifier columns by naming convention."""
    upper = col.upper()
    return any(tag in upper for tag in ("ID_", "_ID", "UUID", "KEY", "HASH"))


def _is_date_like(col: str) -> bool:
    upper = col.upper()
    return any(tag in upper for tag in ("FECHA", "DATE", "DT_", "TIME"))


def _unique_count(col: str, rows: list[dict], limit: int = 100) -> int | None:
    """Count unique values up to `limit`; return None if exceeded."""
    seen: set[str] = set()
    for r in rows:
        v = r.get(col, "").strip()
        if v and v not in seen:
            seen.add(v)
            if len(seen) > limit:
                return None
    return len(seen)


def _auto_detect_roles(
    cols: list[str], rows: list[dict],
    max_cat_cardinality: int = 50,
) -> dict[str, list[str]]:
    """Auto-detect column roles.

    Heuristics:
      - ID-like names → metadata
      - Date-like names (FECHA, DATE, DT_) → metadata
      - High-cardinality strings (>max_cat_cardinality unique) → metadata
      - Numeric values → numeric
      - Everything else → categorical

    Returns dict with keys: categorical, numeric, metadata, skip.
    """
    roles: dict[str, list[str]] = {
        "categorical": [],
        "numeric": [],
        "metadata": [],
        "skip": [],
    }
    for col in cols:
        if _is_id_like(col):
            roles["metadata"].append(col)
        elif _is_date_like(col):
            roles["metadata"].append(col)
        elif _is_numeric(col, rows):
            roles["numeric"].append(col)
        else:
            # Check cardinality before marking categorical
            n_unique = _unique_count(col, rows, max_cat_cardinality)
            if n_unique is None or n_unique > max_cat_cardinality:
                roles["metadata"].append(col)
            else:
                roles["categorical"].append(col)
    return roles


def _build_encoding_maps(
    rows: list[dict],
    cat_cols: list[str],
    num_cols: list[str],
) -> dict:
    """Build category→code mappings and numeric min/max."""
    maps: dict = {"categorical": {}, "numeric": {}}

    for col in cat_cols:
        unique = sorted(
            set(r.get(col, "") for r in rows if r.get(col, "").strip())
        )
        if not unique:
            unique = ["__MISSING__"]
        maps["categorical"][col] = {
            "values": unique,
            "code": {v: i + 1 for i, v in enumerate(unique)},
        }

    for col in num_cols:
        vals = []
        for r in rows:
            try:
                v = float(r.get(col, "") or 0)
                vals.append(v)
            except (ValueError, TypeError):
                vals.append(0.0)
        vals = np.array(vals)
        mn = float(vals.min())
        mx = float(vals.max())
        maps["numeric"][col] = {
            "min": mn,
            "max": mx,
            "range": mx - mn if mx != mn else 1.0,
        }

    return maps


def _compute_feature_dim(maps: dict) -> int:
    cat_dim = sum(len(v["values"]) for v in maps["categorical"].values())
    num_dim = len(maps["numeric"])
    return cat_dim + num_dim


def _build_semantic_projection(
    maps: dict,
    embed_dim: int,
) -> np.ndarray:
    """Build a projection matrix by embedding each category value and
    numeric column via nomic-embed-text, then stacking + PCA whitening."""
    from src.knowledge.embeddings import get_embedding_service

    embedder = get_embedding_service()
    feature_dim = _compute_feature_dim(maps)
    proj = np.zeros((feature_dim, embed_dim))

    idx = 0
    for col_name, col_info in maps["categorical"].items():
        for val in col_info["values"]:
            text = f"[COL] {col_name} [VAL] {val}"
            emb = embedder.embed_one(text)
            proj[idx, :] = emb[:embed_dim]
            idx += 1

    for col_name in maps["numeric"]:
        text = f"[COL] {col_name} [VAL]"
        emb = embedder.embed_one(text)
        proj[idx, :] = emb[:embed_dim]
        idx += 1

    from sklearn.decomposition import PCA
    k = min(embed_dim, feature_dim)
    proj = PCA(n_components=k).fit_transform(proj)
    return proj


def _build_random_projection(feature_dim: int, embed_dim: int) -> np.ndarray:
    rng = np.random.default_rng(42)
    proj = rng.normal(0.0, 1.0 / np.sqrt(feature_dim),
                      size=(feature_dim, embed_dim))
    q, _ = np.linalg.qr(proj)
    return q[:, :embed_dim]


def _generate_sas_code(
    maps: dict,
    meta_cols: list[str],
    embed_dim: int,
) -> str:
    """Generate the SAS include file with format definitions + DATA step."""
    cat_maps = maps["categorical"]
    num_maps = maps["numeric"]

    lines: list[str] = []
    nl = "\n"
    tab = "    "

    # ── Header ──
    lines.append("/* ── Auto-generated SAS embedding code ────────────────── */")
    n_feat = _compute_feature_dim(maps)
    lines.append(f"%let N_FEATURES = {n_feat};")
    lines.append(f"%let N_CAT_COLS = {len(cat_maps)};")
    lines.append(f"%let N_NUM_COLS = {len(num_maps)};")
    lines.append(f"%let EMBED_DIM = {embed_dim};")
    lines.append("")

    # ── Invalue definitions (character→numeric mapping) ──
    lines.append("proc format;")
    for col_name, col_info in cat_maps.items():
        fmt_name = f"_{col_name.replace(' ', '_')}_"
        lines.append(f"  invalue {fmt_name}")
        for val in col_info["values"]:
            code = col_info["code"][val]
            safe_val = val.replace("'", "''")
            lines.append(f"    '{safe_val}' = {code}")
        lines.append("    other = 0;")
        lines.append("")
    lines.append("run;")
    lines.append("")

    # ── Feature construction DATA step ──
    lines.append("data _features;")
    lines.append("    set cycles;")
    lines.append("")
    lines.append(f"    array _f[{n_feat}] _temporary_;")
    lines.append(f"    do _i = 1 to {n_feat}; _f[_i] = 0; end;")
    lines.append("")

    # One-hot encoding
    feat_idx = 0
    for col_name, col_info in cat_maps.items():
        n_cats = len(col_info["values"])
        fmt_name = f"_{col_name.replace(' ', '_')}_"
        safe_col = col_name.replace(" ", "_")
        lines.append(
            f"    _code = input({safe_col}, {fmt_name}.);"
        )
        lines.append(f"    if _code > 0 then _f[{feat_idx + 1} + _code - 1] = 1;")
        feat_idx += n_cats

    # Numeric normalization
    for col_name, col_info in num_maps.items():
        mn = col_info["min"]
        rng = col_info["range"]
        safe_col = col_name.replace(" ", "_")
        if rng > 0:
            lines.append(
                f"    _f[{feat_idx + 1}] = "
                f"(input({safe_col}, best32.) - {mn:.8g}) / {rng:.8g};"
            )
        else:
            lines.append(f"    _f[{feat_idx + 1}] = 0;")
        feat_idx += 1

    lines.append("")
    lines.append(f"    array _out[{n_feat}] F1-F{n_feat};")
    lines.append(f"    do _i = 1 to {n_feat}; _out[_i] = _f[_i]; end;")
    lines.append("")
    lines.append("    _ROWID_ = _N_;")
    lines.append("")
    # Progress log
    lines.append("    if mod(_N_, 500) = 0 then")
    lines.append("        put 'NOTE: [TabBERT-SAS] Row ' _N_ ' built at ' datetime20.;")
    lines.append(f"    keep F1-F{n_feat} _ROWID_;")
    lines.append("run;")
    lines.append("")

    # ── Metadata snapshot ──
    lines.append("data _meta;")
    lines.append("    set cycles;")
    lines.append("    _ROWID_ = _N_;")
    lines.append("    keep _ROWID_ _all_;")
    lines.append("run;")
    lines.append("")

    # ── PROC IML code ──
    lines.append("proc iml;")
    lines.append("    use proj_raw;")
    lines.append("    read all var _num_ into W;")
    lines.append("    close proj_raw;")
    lines.append("")
    lines.append("    use _features;")
    lines.append("    read all var _num_ into X_all;")
    lines.append("    close _features;")
    lines.append("")
    lines.append("    n = nrow(X_all);")
    lines.append("    f_dim = ncol(X_all) - 1;")
    lines.append("    X = X_all[, 1:f_dim];")
    lines.append("    row_ids = X_all[, ncol(X_all)];")
    lines.append("    e_dim = ncol(W);")
    lines.append("")
    lines.append('    print "TabBERT-SAS: " n "cases, " f_dim "features -> " e_dim "embeddings";')
    lines.append("    print '---';")
    lines.append("")
    lines.append("    n_batches = ceil(n / 5000);")
    lines.append("    do b = 1 to n_batches;")
    lines.append("        start = (b-1) * 5000 + 1;")
    lines.append("        end_  = min(b * 5000, n);")
    lines.append("        X_batch  = X[start:end_, ];")
    lines.append("        ids_batch = row_ids[start:end_, ];")
    lines.append("        E_batch = X_batch * W;")
    lines.append("        out_batch = ids_batch || E_batch;")
    lines.append("        if b = 1 then E = out_batch;")
    lines.append("        else E = E // out_batch;")
    lines.append("        pct = round(end_ / n * 100, 1);")
    lines.append('        print "Batch " b "of " n_batches ": " end_ "of " n " (" pct "%)";')
    lines.append("    end;")
    lines.append("    print '---';")
    lines.append("")

    # Build column names dynamically (IML-native: no cats() needed)
    lines.append("    indices = 1:e_dim;")
    lines.append("    emb_names = 'EMB_' + strip(char(indices));")
    lines.append("    col_names = 'ROW_ID' || emb_names;")
    lines.append("")
    lines.append("    create _embeddings from E [colname=col_names];")
    lines.append("    append from E;")
    lines.append("    close _embeddings;")
    lines.append("    print 'TabBERT-SAS: Done';")
    lines.append("quit;")
    lines.append("")

    # ── Merge + export ──
    lines.append("proc sql;")
    lines.append("    create table sas_embeddings_full as")
    lines.append("        select e.*, m.*")
    lines.append("        from _embeddings e")
    lines.append("        left join _meta m on e.ROW_ID = m._ROWID_")
    lines.append("        order by e.ROW_ID;")
    lines.append("    drop table _features, _meta, _embeddings;")
    lines.append("quit;")
    lines.append("")

    return nl.join(lines)


def _write_excel(
    output_dir: Path,
    projection: np.ndarray,
    maps: dict,
    meta_cols: list[str],
    embed_dim: int,
) -> None:
    """Write sas_weights.xlsx with 4 sheets for SAS Excel import."""
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Projection
    ws1 = wb.active
    ws1.title = "Projection"
    ws1.append(["FEAT"] + [f"D{j+1}" for j in range(embed_dim)])
    for i in range(projection.shape[0]):
        ws1.append([f"F{i+1}"] + [float(projection[i, j]) for j in range(embed_dim)])

    # Sheet 2: CatMaps
    ws2 = wb.create_sheet("CatMaps")
    ws2.append(["COLUMN", "VALUE", "CODE"])
    for col_name, col_info in maps["categorical"].items():
        for val in col_info["values"]:
            ws2.append([col_name, val, col_info["code"][val]])

    # Sheet 3: NumMaps
    ws3 = wb.create_sheet("NumMaps")
    ws3.append(["COLUMN", "MIN", "MAX", "RANGE"])
    for col_name, info in maps["numeric"].items():
        ws3.append([col_name, info["min"], info["max"], info["range"]])

    # Sheet 4: Config
    ws4 = wb.create_sheet("Config")
    ws4.append(["KEY", "VALUE"])
    ws4.append(["categorical_cols", ", ".join(maps["categorical"].keys())])
    ws4.append(["numeric_cols", ", ".join(maps["numeric"].keys())])
    ws4.append(["metadata_cols", ", ".join(meta_cols)])
    ws4.append(["feature_dim", projection.shape[0]])
    ws4.append(["embed_dim", embed_dim])

    xlsx_path = output_dir / "sas_weights.xlsx"
    wb.save(xlsx_path)
    size_kb = xlsx_path.stat().st_size / 1024
    print(f"  Wrote {xlsx_path}  ({size_kb:.0f} KB)")


def _generate_python_config(maps: dict, meta_cols: list[str]) -> dict:
    """Generate a small JSON config for the Python simulation script."""
    cat_info = {}
    for col, info in maps["categorical"].items():
        cat_info[col] = info["values"]
    num_info = {}
    for col, info in maps["numeric"].items():
        num_info[col] = {"min": info["min"], "range": info["range"]}
    return {
        "categorical_cols": list(maps["categorical"].keys()),
        "numeric_cols": list(maps["numeric"].keys()),
        "metadata_cols": meta_cols,
        "cat_values": cat_info,
        "num_params": num_info,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export embedding weights + SAS code for any CSV"
    )
    parser.add_argument("--csv", type=Path, required=True,
                        help="Input CSV file (any table)")
    parser.add_argument("--embed-dim", type=int, default=50,
                        help="Embedding dimension (default: 50)")
    parser.add_argument("--method", type=str, default="semantic",
                        choices=["random", "semantic"])
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    # Manual column role overrides
    parser.add_argument("--cat-cols", nargs="*", default=None)
    parser.add_argument("--num-cols", nargs="*", default=None)
    parser.add_argument("--meta-cols", nargs="*", default=None)
    parser.add_argument("--skip-cols", nargs="*", default=None)
    parser.add_argument("--max-cat-cardinality", type=int, default=50,
                        help="Max unique values for categorical auto-detect (default: 50)")
    args = parser.parse_args()

    if not args.csv.exists():
        print(f"CSV not found: {args.csv}")
        sys.exit(1)

    cols, rows = _load_csv(args.csv)
    print(f"Loaded {len(rows)} rows × {len(cols)} cols from {args.csv}")

    # Detect or use explicit roles
    roles = _auto_detect_roles(cols, rows, max_cat_cardinality=args.max_cat_cardinality)
    if args.cat_cols is not None:
        roles["categorical"] = [c for c in args.cat_cols if c in cols]
    if args.num_cols is not None:
        roles["numeric"] = [c for c in args.num_cols if c in cols]
    if args.meta_cols is not None:
        roles["metadata"] = [c for c in args.meta_cols if c in cols]
    if args.skip_cols is not None:
        roles["skip"] = [c for c in args.skip_cols if c in cols]
        roles["categorical"] = [c for c in roles["categorical"] if c not in roles["skip"]]
        roles["numeric"] = [c for c in roles["numeric"] if c not in roles["skip"]]
        roles["metadata"] = [c for c in roles["metadata"] if c not in roles["skip"]]
    # Metadata: anything not cat/num/skip
    used = set(roles["categorical"]) | set(roles["numeric"]) | set(roles["skip"])
    roles["metadata"] = list(set(roles["metadata"]) | (set(cols) - used))

    cat_cols = roles["categorical"]
    num_cols = roles["numeric"]
    meta_cols = roles["metadata"]

    print(f"Roles: {len(cat_cols)} categorical, {len(num_cols)} numeric, "
          f"{len(meta_cols)} metadata, {len(roles['skip'])} skipped")

    if not cat_cols and not num_cols:
        print("ERROR: no categorical or numeric columns found")
        sys.exit(1)

    # Build encoding maps
    maps = _build_encoding_maps(rows, cat_cols, num_cols)
    feature_dim = _compute_feature_dim(maps)
    print(f"Feature vector: {feature_dim} dims "
          f"({sum(len(v['values']) for v in maps['categorical'].values())} cat + "
          f"{len(num_cols)} num)")

    # Build projection matrix
    if args.method == "semantic":
        print("Building semantic projection via nomic-embed-text...")
        projection = _build_semantic_projection(maps, args.embed_dim)
    else:
        print(f"Building random projection ({feature_dim}→{args.embed_dim})...")
        projection = _build_random_projection(feature_dim, args.embed_dim)

    embed_dim = projection.shape[1]
    projection = projection[:, :embed_dim]
    print(f"Projection matrix: {projection.shape[0]}×{projection.shape[1]}")

    # Write output files
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Projection matrix CSV
    proj_path = args.output_dir / "sas_projection.csv"
    with open(proj_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["FEAT"] + [f"D{j+1}" for j in range(embed_dim)])
        for i in range(projection.shape[0]):
            w.writerow([f"F{i+1}"] + [f"{v:.10f}" for v in projection[i]])
    print(f"  Wrote {proj_path}")

    # 2. Categorical maps CSV
    cat_path = args.output_dir / "sas_cat_maps.csv"
    with open(cat_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["COLUMN", "VALUE", "CODE"])
        for col_name, col_info in maps["categorical"].items():
            for val in col_info["values"]:
                w.writerow([col_name, val, col_info["code"][val]])
    print(f"  Wrote {cat_path}")

    # 3. Numeric maps CSV
    num_path = args.output_dir / "sas_num_maps.csv"
    with open(num_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["COLUMN", "MIN", "MAX", "RANGE"])
        for col_name, info in maps["numeric"].items():
            w.writerow([col_name, info["min"], info["max"], info["range"]])
    print(f"  Wrote {num_path}")

    # 4. SAS generated code
    sas_code = _generate_sas_code(maps, meta_cols, embed_dim)
    sas_path = args.output_dir / "sas_generated.sas"
    sas_path.write_text(sas_code)
    print(f"  Wrote {sas_path}  ({len(sas_code)} chars)")

    # 5. Python config (JSON)
    py_cfg = _generate_python_config(maps, meta_cols)
    cfg_path = args.output_dir / "sas_config.json"
    import json
    cfg_path.write_text(json.dumps(py_cfg, indent=2, ensure_ascii=False))
    print(f"  Wrote {cfg_path}")

    # 6. Excel workbook for SAS users
    _write_excel(args.output_dir, projection, maps, meta_cols, embed_dim)

    print(f"\nDone. Generated {feature_dim} features → {embed_dim} embeddings.")
    print(f"  SAS:     sas scripts/tabular_embeddings.sas")
    print(f"  Python:  python scripts/simulate_sas_embeddings.py")


if __name__ == "__main__":
    main()
