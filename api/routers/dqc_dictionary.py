"""Dictionary intelligence for the DQC generator.

Turns a raw Excel workbook into LLM-ready dictionary context, with the
model doing the judgment calls the old hard-coded parser could not:

* **Sheet selection** — a workbook may carry several sheets (notes,
  changelog, the actual field dictionary). Heuristics score every sheet;
  when the answer is not obvious, ONE small LLM call proposes the sheet
  and, failing confidence, a clarifying question the chat UI shows the
  user (who answers by clicking a sheet name).
* **Column mapping** — which workbook column holds the field name, the
  type, the description, the formula, the regulatory reference. The LLM
  proposes the mapping from headers + a few sample rows; the UI can show
  and override it. Hard-coded header matching remains as the fallback.
* **Format inference** — fields whose type cell is empty get a proposed
  format (REAL/INTEGER/TEXT/DATE/YYYYMM + nullability) from name and
  description, clearly marked ``(inferido)`` in the prompt context.

Context-window discipline (the local GGUF backend defaults to an 8k
window): every LLM call in this module and in the generation path is a
**fresh, stateless agent** (one ``chat_json`` call, no shared
conversation), receives only the rows it needs, and is capped by
``PROMPT_CHAR_BUDGET``. Generation batches instructions and sends each
batch only the dictionary fields relevant to it.
"""

from __future__ import annotations

import io
import json
import logging
import re
import unicodedata
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# ── budgets (chars ≈ tokens×4; sized for an 8k-token local context) ─────────
PROMPT_CHAR_BUDGET = 12_000     # max chars of dictionary text per LLM call
MAX_FIELDS_PER_CALL = 60        # hard cap on fields sent to one agent
SAMPLE_ROWS = 4                 # sample rows per sheet shown to the mapper
CELL_TRUNC = 60                 # cell preview length

ROLE_KEYS = ("field", "type", "description", "nullable", "formula", "reg_ref")

_HEADER_HINTS = {
    "field": ("field", "campo", "columna", "nombre", "name", "variable"),
    "type": ("type", "tipo", "datatype", "formato"),
    "description": ("description", "descripcion", "descripción", "definicion",
                    "definición", "desc"),
    "nullable": ("null", "nulo", "nullable", "obligatorio", "opcional"),
    "formula": ("formula", "fórmula", "derivacion", "derivación", "calculo",
                "cálculo"),
    "reg_ref": ("reg ref", "reg_ref", "referencia", "regulacion", "regulación",
                "normativa", "articulo", "artículo"),
}


def _norm(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.strip().lower()


# ── workbook inspection ──────────────────────────────────────────────────────

@dataclass
class SheetInfo:
    name: str
    n_rows: int
    headers: list[str]
    samples: list[list[str]]
    score: int = 0


@dataclass
class Inspection:
    sheets: list[SheetInfo] = field(default_factory=list)

    @property
    def best(self) -> SheetInfo | None:
        ranked = sorted(self.sheets, key=lambda s: s.score, reverse=True)
        return ranked[0] if ranked else None

    @property
    def ambiguous(self) -> bool:
        """True when we should ask instead of guessing."""
        ranked = sorted(self.sheets, key=lambda s: s.score, reverse=True)
        if not ranked:
            return True
        if len(ranked) == 1:
            return ranked[0].score < 2
        return ranked[0].score < 2 or ranked[0].score == ranked[1].score


def _heuristic_mapping(headers: list[str]) -> dict[str, str | None]:
    lower = {_norm(h): h for h in headers if h}
    out: dict[str, str | None] = {}
    for role, hints in _HEADER_HINTS.items():
        out[role] = None
        for hint in hints:
            if hint in lower:
                out[role] = lower[hint]
                break
    return out


def _score_sheet(info: SheetInfo) -> int:
    mapping = _heuristic_mapping(info.headers)
    score = 0
    if mapping["field"]:
        score += 2
    if mapping["type"]:
        score += 1
    if mapping["description"]:
        score += 1
    if info.n_rows >= 5:
        score += 1
    if info.n_rows == 0:
        score = 0
    return score


def inspect_workbook(data: bytes) -> Inspection:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    inspection = Inspection()
    for ws in wb.worksheets:
        headers: list[str] = []
        samples: list[list[str]] = []
        n_rows = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            cells = [str(v or "").strip()[:CELL_TRUNC] for v in row]
            if not headers:
                headers = cells
                continue
            n_rows += 1
            if len(samples) < SAMPLE_ROWS:
                samples.append(cells)
        info = SheetInfo(name=ws.title, n_rows=n_rows,
                         headers=headers, samples=samples)
        info.score = _score_sheet(info)
        inspection.sheets.append(info)
    wb.close()
    return inspection


# ── LLM mapping proposal (one fresh agent, tiny context) ─────────────────────

_MAPPER_SYSTEM = """\
Eres un asistente que identifica diccionarios de campos en libros Excel.
Recibes las hojas de un libro (cabeceras + filas de ejemplo). Devuelve SOLO JSON:
{
  "sheet": "<hoja que contiene el diccionario de campos>",
  "column_mapping": {
    "field": "<cabecera con el NOMBRE del campo>",
    "type": "<cabecera con el tipo/formato o null>",
    "description": "<cabecera con la descripción o null>",
    "nullable": "<cabecera de nulabilidad o null>",
    "formula": "<cabecera con fórmulas de derivación o null>",
    "reg_ref": "<cabecera con referencia regulatoria o null>"
  },
  "confidence": <0.0-1.0>,
  "question": "<pregunta breve para el usuario SOLO si confidence < 0.6, si no null>"
}
Usa exactamente los nombres de hoja y cabecera recibidos."""


def propose_mapping(inspection: Inspection, client) -> dict:
    """Ask ONE stateless agent to pick the sheet + column mapping.

    Falls back to heuristics if the model is unavailable or answers
    garbage. Returns::

        {sheet, column_mapping, confidence, source, question, options}
    """
    heur_sheet = inspection.best
    heuristic = {
        "sheet": heur_sheet.name if heur_sheet else None,
        "column_mapping": _heuristic_mapping(heur_sheet.headers) if heur_sheet else {},
        "confidence": 0.5 if heur_sheet and heur_sheet.score >= 3 else 0.3,
        "source": "heuristic",
        "question": None,
        "options": [s.name for s in inspection.sheets],
    }

    parts = []
    for s in inspection.sheets:
        parts.append(f"HOJA {s.name!r} ({s.n_rows} filas)")
        parts.append(f"  cabeceras: {s.headers}")
        for row in s.samples[:2]:
            parts.append(f"  ejemplo: {row}")
    user = "\n".join(parts)[:PROMPT_CHAR_BUDGET]

    try:
        result = client.chat_json(system=_MAPPER_SYSTEM, user=user,
                                  max_tokens=512)
    except Exception as exc:  # noqa: BLE001 — degrade to heuristics
        logger.warning("mapping proposal failed: %s", exc)
        return heuristic

    sheet_names = {s.name for s in inspection.sheets}
    sheet = result.get("sheet") if isinstance(result, dict) else None
    if sheet not in sheet_names:
        return heuristic

    raw_mapping = result.get("column_mapping") or {}
    headers = {h for s in inspection.sheets if s.name == sheet for h in s.headers}
    mapping = {}
    for role in ROLE_KEYS:
        value = raw_mapping.get(role)
        mapping[role] = value if value in headers else None
    if not mapping["field"]:
        # a dictionary without a field-name column is no mapping at all
        return heuristic

    try:
        confidence = float(result.get("confidence", 0.5))
    except (TypeError, ValueError):
        confidence = 0.5
    question = result.get("question") or None
    if confidence < 0.6 and not question:
        question = "¿Qué hoja del Excel contiene el diccionario de campos?"

    return {
        "sheet": sheet, "column_mapping": mapping,
        "confidence": confidence, "source": "llm",
        "question": question if confidence < 0.6 else None,
        "options": [s.name for s in inspection.sheets],
    }


# ── dictionary parsing with an explicit mapping ──────────────────────────────

@dataclass
class FieldEntry:
    name: str
    type: str = ""
    description: str = ""
    nullable: str = ""
    formula: str = ""
    reg_ref: str = ""
    type_inferred: bool = False

    def to_line(self) -> str:
        parts = [f"- {self.name}"]
        if self.type:
            suffix = " (inferido)" if self.type_inferred else ""
            parts.append(f"type={self.type}{suffix}")
        if self.nullable:
            parts.append(f"null={self.nullable}")
        if self.description:
            parts.append(f"desc={self.description}")
        if self.formula:
            parts.append(f"formula={self.formula}")
        if self.reg_ref:
            parts.append(f"reg={self.reg_ref}")
        return ", ".join(parts)


def parse_dictionary(data: bytes, sheet: str | None = None,
                     mapping: dict | None = None) -> list[FieldEntry]:
    """Read the field dictionary using an explicit sheet + column mapping
    (falling back to heuristics for anything unmapped)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    if ws is None:
        wb.close()
        return []

    rows_iter = ws.iter_rows(values_only=True)
    header_row = None
    for row in rows_iter:
        if row and any(v is not None and str(v).strip() for v in row):
            header_row = row
            break
    if not header_row:
        wb.close()
        return []
    headers = [str(c or "").strip() for c in header_row]

    resolved = _heuristic_mapping(headers)
    if mapping:
        for role in ROLE_KEYS:
            if mapping.get(role) in headers:
                resolved[role] = mapping[role]

    def col_idx(role: str) -> int | None:
        name = resolved.get(role)
        return headers.index(name) if name in headers else None

    idx = {role: col_idx(role) for role in ROLE_KEYS}
    fields: list[FieldEntry] = []
    for row in rows_iter:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(role: str) -> str:
            i = idx[role]
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        name = cell("field")
        if not name:
            continue
        fields.append(FieldEntry(
            name=name, type=cell("type"), description=cell("description"),
            nullable=cell("nullable"), formula=cell("formula"),
            reg_ref=cell("reg_ref")))
    wb.close()
    return fields


# ── format inference (fresh agent per batch of untyped fields) ───────────────

_FORMAT_SYSTEM = """\
Eres un experto en modelos de datos bancarios. Recibes campos SIN tipo
declarado (nombre + descripción). Infiere el formato de cada uno.
Devuelve SOLO JSON: {"formats": {"<CAMPO>": {"tipo": "REAL|INTEGER|TEXT|DATE|YYYYMM",
"nulo": "si|no"}}}. Incluye TODOS los campos recibidos."""


def infer_missing_formats(fields: list[FieldEntry], client,
                          batch_size: int = 25) -> int:
    """Fill empty ``type`` values via stateless LLM calls (one fresh agent
    per batch — bounded context). Returns how many were inferred."""
    untyped = [f for f in fields if not f.type]
    inferred = 0
    for start in range(0, len(untyped), batch_size):
        batch = untyped[start:start + batch_size]
        user = "\n".join(
            f"- {f.name}: {f.description or '(sin descripción)'}"
            for f in batch)[:PROMPT_CHAR_BUDGET]
        try:
            result = client.chat_json(system=_FORMAT_SYSTEM, user=user,
                                      max_tokens=1024)
        except Exception as exc:  # noqa: BLE001
            logger.warning("format inference failed: %s", exc)
            break
        formats = result.get("formats") if isinstance(result, dict) else None
        if not isinstance(formats, dict):
            continue
        lookup = {_norm(k): v for k, v in formats.items()}
        for entry in batch:
            info = lookup.get(_norm(entry.name))
            if isinstance(info, dict) and info.get("tipo"):
                entry.type = str(info["tipo"]).upper()
                entry.type_inferred = True
                if info.get("nulo") and not entry.nullable:
                    entry.nullable = str(info["nulo"])
                inferred += 1
    return inferred


# ── relevance selection + batching (context-window discipline) ───────────────

_WORD_RE = re.compile(r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ_0-9]{3,}")

# hybrid weights when a semantic (embedding) channel is available; lexical is
# kept dominant so an outright-named field always outranks a merely-similar one
_LEX_WEIGHT = 0.6
_SEM_WEIGHT = 0.4


def _lexical_score(entry: FieldEntry, words: set[str]) -> float:
    name_n = _norm(entry.name)
    score = 0
    if name_n in words:
        score += 100                     # named outright
    score += sum(3 for part in name_n.split("_") if part in words)
    desc_words = {_norm(w) for w in _WORD_RE.findall(entry.description)}
    score += len(words & desc_words)
    formula_words = {_norm(w) for w in _WORD_RE.findall(entry.formula)}
    score += 2 * len(words & formula_words)
    return float(score)


def _cosine(a: list[float], b: list[float]) -> float:
    dot = num = den_a = den_b = 0.0
    for x, y in zip(a, b):
        dot += x * y
        den_a += x * x
        den_b += y * y
    if den_a == 0.0 or den_b == 0.0:
        return 0.0
    return dot / ((den_a ** 0.5) * (den_b ** 0.5))


def _minmax(values: list[float]) -> list[float]:
    lo, hi = min(values), max(values)
    if hi - lo < 1e-12:
        return [0.0] * len(values)
    return [(v - lo) / (hi - lo) for v in values]


def select_relevant_fields(fields: list[FieldEntry], instructions: list[str],
                           cap: int = MAX_FIELDS_PER_CALL,
                           embedder=None) -> list[FieldEntry]:
    """Fields worth sending to the agent for THIS instruction batch.

    Tier-1 schema-linking filter (MARS-SQL style): rank every field and keep
    the top ``cap``. Lexical scoring (exact name, name parts, description /
    formula word overlap) is always applied; when ``embedder`` is provided
    and actually produces vectors, a semantic channel (cosine similarity of
    the rule against each field's ``name. description``) is blended in so
    fields that are relevant but share no literal words are still surfaced.
    Falls back to pure lexical when there is no embedder or it is unavailable
    (zero vectors), so behaviour is unchanged unless embeddings are on.
    """
    if len(fields) <= cap:
        return fields
    text = " ".join(instructions)
    words = {_norm(w) for w in _WORD_RE.findall(text)}
    lex = [_lexical_score(e, words) for e in fields]

    sem: list[float] | None = None
    if embedder is not None:
        try:
            field_texts = [f"{e.name}. {e.description}".strip() for e in fields]
            vecs = embedder.embed([text] + field_texts)
            query_vec = vecs[0]
            if any(query_vec):   # non-zero ⇒ embeddings really ran (not stub)
                sem = [_cosine(query_vec, fv) for fv in vecs[1:]]
        except Exception as exc:  # noqa: BLE001 — semantic channel is best-effort
            logger.warning("semantic field filter failed: %s", exc)
            sem = None

    if sem is None:
        combined = lex
    else:
        lex_n, sem_n = _minmax(lex), _minmax(sem)
        combined = [_LEX_WEIGHT * l + _SEM_WEIGHT * s
                    for l, s in zip(lex_n, sem_n)]

    order = sorted(range(len(fields)), key=lambda i: (-combined[i], i))
    chosen = {id(fields[i]) for i in order[:cap]}
    # keep original dictionary order for the prompt
    return [f for f in fields if id(f) in chosen]


def fields_to_text(fields: list[FieldEntry],
                   budget: int = PROMPT_CHAR_BUDGET) -> tuple[str, int]:
    """Render field lines under the char budget; returns (text, used)."""
    lines: list[str] = []
    used = 0
    count = 0
    for entry in fields:
        line = entry.to_line()
        if used + len(line) + 1 > budget:
            break
        lines.append(line)
        used += len(line) + 1
        count += 1
    return "\n".join(lines), count


# ── uploaded natural-language test lists ─────────────────────────────────────

_BULLET_RE = re.compile(r"^\s*(?:[-*•·]|\d+[.)]|[a-zA-Z][.)])\s+")


def _clean_instruction(line: str) -> str:
    """One rule per line; leading bullets/numbering stripped."""
    return _BULLET_RE.sub("", line).strip()


def read_instructions_upload(filename: str, data: bytes) -> list[str]:
    """Extract natural-language DQC rules from an uploaded list.

    Supported: .txt / .md (one rule per line), .csv (first non-empty cell
    per row), .xlsx (first sheet, first text column; a header row that
    looks like a title — 'test', 'regla', 'instruccion', 'dqc',
    'descripcion' — is skipped).
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                    data_only=True)
        ws = wb.active
        rules: list[str] = []
        if ws is not None:
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                cell = next((str(v).strip() for v in row
                             if v is not None and str(v).strip()), "")
                if cell:
                    rules.append(cell)
        wb.close()
    elif name.endswith(".csv"):
        text = data.decode("utf-8", errors="replace")
        rules = []
        for line in text.splitlines():
            cell = next((c.strip().strip('"') for c in line.split(",")
                         if c.strip().strip('"')), "")
            if cell:
                rules.append(cell)
    else:  # .txt / .md / anything text-like
        text = data.decode("utf-8", errors="replace")
        rules = [ln for ln in text.splitlines() if ln.strip()]

    cleaned = [c for c in (_clean_instruction(r) for r in rules) if c]
    # drop a lone header/title row (common in xlsx/csv lists)
    if cleaned and _norm(cleaned[0]) in (
            "test", "tests", "regla", "reglas", "instruccion", "instrucciones",
            "dqc", "dqcs", "descripcion", "control", "controles", "check",
            "checks"):
        cleaned = cleaned[1:]
    return cleaned


def plan_batches(lines: list[str], batch_size: int) -> list[list[str]]:
    batch_size = max(1, batch_size)
    return [lines[i:i + batch_size] for i in range(0, len(lines), batch_size)]
