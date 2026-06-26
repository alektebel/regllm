"""Agentic Q&A router.

Endpoints
---------
POST /agent/ask                    SSE stream of agent events for one question
POST /agent/sas/upload             Upload .sas/.egp files into data/sas/sessions/{session_id}/
POST /agent/docs/reindex           Rebuild the BM25 docs index
GET  /agent/status                 Counts + active LLM backend/model
"""

from __future__ import annotations

import json
import logging
import shutil
from pathlib import Path
from typing import AsyncIterator

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/agent", tags=["agent"])

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_SAS_ROOT = _PROJECT_ROOT / "data" / "sas"
_DOCS_ROOT = _PROJECT_ROOT / "data" / "docs"
_INDEX_PATH = _DOCS_ROOT / "index.json"


# ── Request/response models ──────────────────────────────────────────────────


class AskRequest(BaseModel):
    question: str = Field(..., min_length=3)
    max_iters: int = 15
    temperature: float = 0.1
    session_id: str | None = None
    sas_session_id: str | None = "debug_lgd"  # SAS code folder under data/sas/sessions/
    context: list[dict] | None = None  # injected system messages (e.g. upload notifications)
    model: str | None = None           # override Ollama model (e.g. "llama3.2")
    tool_names: list[str] | None = None  # restrict to subset of tools


# ── /agent/ask  (SSE) ────────────────────────────────────────────────────────


def _sse_event(event: dict) -> bytes:
    """Format one SSE message. Single-line ``data:`` is the simplest path."""
    return ("data: " + json.dumps(event, default=str) + "\n\n").encode("utf-8")


async def _stream(
    question: str,
    *,
    max_iters: int,
    temperature: float,
    session_id: str | None = None,
    sas_session_id: str | None = "debug_lgd",
    context: list[dict] | None = None,
    model: str | None = None,
    tool_names: list[str] | None = None,
) -> AsyncIterator[bytes]:
    from src.agent import SASDiffAgent

    agent = SASDiffAgent(
        max_iters=max_iters,
        temperature=temperature,
        session_id=session_id,
        sas_session_id=sas_session_id,
        model=model,
        tool_names=tool_names,
    )

    # Build effective question with any injected context
    effective_question = question
    if context:
        ctx_text = "\n".join(
            f"[{c.get('role', 'system')}] {c.get('content', '')}"
            for c in context
        )
        effective_question = f"{ctx_text}\n\nUser question: {question}"

    # Save session state if session_id provided
    if session_id:
        _save_session(session_id, question, context)

    try:
        async for ev in agent.run(effective_question):
            yield _sse_event(ev.to_dict())
    except Exception as e:
        logger.exception("Agent run failed")
        yield _sse_event({"type": "error", "stage": "run", "error": str(e)})
    yield _sse_event({"type": "done"})


@router.post("/ask")
async def ask(req: AskRequest) -> StreamingResponse:
    return StreamingResponse(
        _stream(
            req.question,
            max_iters=req.max_iters,
            temperature=req.temperature,
            session_id=req.session_id,
            sas_session_id=req.sas_session_id,
            context=req.context,
            model=req.model,
            tool_names=req.tool_names,
        ),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── Session state ────────────────────────────────────────────────────────────

_SESSIONS_ROOT = _PROJECT_ROOT / "data" / "sessions"
_UPLOADS_ROOT = _PROJECT_ROOT / "data" / "uploads"


def _save_session(session_id: str, question: str, context: list[dict] | None) -> None:
    _SESSIONS_ROOT.mkdir(parents=True, exist_ok=True)
    session_file = _SESSIONS_ROOT / f"{session_id}.json"
    data: dict = {}
    if session_file.exists():
        data = json.loads(session_file.read_text(encoding="utf-8"))
    if "questions" not in data:
        data["questions"] = []
    data["questions"].append({"question": question, "context": context})
    session_file.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


# ── /agent/data/upload ──────────────────────────────────────────────────────


@router.post("/data/upload")
async def upload_data(
    files: list[UploadFile] = File(...),
    session_id: str = Form("default"),
) -> dict:
    """Upload CSV/XLSX data files for the agent to analyze."""
    target_dir = _UPLOADS_ROOT / session_id
    target_dir.mkdir(parents=True, exist_ok=True)
    results: list[dict] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in {".csv", ".xlsx", ".xls"}:
            raise HTTPException(400, f"Only .csv and .xlsx files allowed (got {ext!r})")
        target = target_dir / Path(f.filename or "").name
        with target.open("wb") as out:
            shutil.copyfileobj(f.file, out)
        # Parse and return preview
        file_info: dict = {"file": target.name, "path": f"{session_id}/{target.name}"}
        try:
            if ext == ".csv":
                import csv as csv_mod
                with target.open(encoding="utf-8") as cf:
                    reader = csv_mod.DictReader(cf)
                    rows = list(reader)
                file_info["columns"] = list(rows[0].keys()) if rows else []
                file_info["row_count"] = len(rows)
                file_info["preview"] = rows[:5]
            elif ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
                ws = wb.active
                if ws is not None:
                    header = [str(c.value or f"col_{i}") for i, c in enumerate(next(ws.iter_rows(max_row=1)))]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(header, row)))
                    file_info["columns"] = header
                    file_info["row_count"] = len(rows)
                    file_info["preview"] = rows[:5]
                wb.close()
        except Exception as e:
            file_info["parse_error"] = str(e)
        results.append(file_info)
    return {"session_id": session_id, "files": results}


# ── /agent/sas/upload ────────────────────────────────────────────────────────

_SAS_SESSIONS_ROOT = _SAS_ROOT / "sessions"


def _sanitize_filename(name: str) -> str:
    """Convert a task name to a safe filename component."""
    import re
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"[\s]+", "_", name).strip("_")
    return name[:60] or "unnamed"


def _extract_egp(
    upload: UploadFile, target_dir: Path,
) -> list[dict]:
    """Parse an .egp upload into individual .sas files + metadata."""
    import tempfile
    from src.sas_parser import SASParser

    blocks_metadata: list[dict] = []
    egp_stem = Path(upload.filename or "project").stem

    with tempfile.NamedTemporaryFile(suffix=".egp", delete=False) as tmp:
        shutil.copyfileobj(upload.file, tmp)
        tmp_path = Path(tmp.name)

    try:
        parser = SASParser()
        blocks = parser.parse(tmp_path)
        if not blocks:
            raise HTTPException(
                422,
                f"No SAS code blocks found in {upload.filename}. "
                "The .egp file may be empty or in an unsupported format.",
            )
        for i, block in enumerate(blocks):
            safe_name = _sanitize_filename(block.task_name or f"block_{i}")
            fname = f"{egp_stem}_{safe_name}.sas"
            out_path = target_dir / fname
            out_path.write_text(block.code, encoding="utf-8")
            blocks_metadata.append({
                "filename": fname,
                "task_name": block.task_name or f"block_{i}",
                "block_type": block.block_type,
                "order": i,
                "line_count": len(block.code.splitlines()),
            })
    finally:
        tmp_path.unlink(missing_ok=True)

    return blocks_metadata


def _write_manifest(
    target_dir: Path, source_file: str, blocks_metadata: list[dict],
) -> None:
    """Write a project manifest alongside extracted .sas files."""
    from datetime import datetime
    manifest = {
        "source_file": source_file,
        "extracted_at": datetime.now().isoformat(),
        "blocks": blocks_metadata,
    }
    (target_dir / "_project_manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8",
    )


@router.post("/sas/upload")
async def upload_sas(
    files: list[UploadFile] = File(...),
    session_id: str = Form("default"),
) -> dict:
    """Upload .sas or .egp files into data/sas/sessions/{session_id}/."""
    target_dir = _SAS_SESSIONS_ROOT / session_id
    target_dir.mkdir(parents=True, exist_ok=True)
    saved: list[str] = []
    all_metadata: list[dict] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in {".sas", ".egp"}:
            raise HTTPException(400, f"Only .sas and .egp files allowed (got {ext!r})")
        if ext == ".egp":
            metadata = _extract_egp(f, target_dir)
            all_metadata.extend(metadata)
            saved.extend(m["filename"] for m in metadata)
            _write_manifest(target_dir, f.filename or "project.egp", metadata)
        else:
            target = target_dir / Path(f.filename or "").name
            with target.open("wb") as out:
                shutil.copyfileobj(f.file, out)
            saved.append(target.name)
    return {
        "session_id": session_id,
        "saved": saved,
        "blocks_metadata": all_metadata,
        "folder": str(target_dir),
    }


@router.post("/sas/upload-egp")
async def upload_egp(
    file: UploadFile = File(...),
    session_id: str = Form("default"),
) -> dict:
    """Upload a .egp (SAS Enterprise Guide project) file.

    Extracts all SAS code blocks from the project and saves them as
    individual .sas files into data/sas/sessions/{session_id}/.
    """
    ext = Path(file.filename or "").suffix.lower()
    if ext != ".egp":
        raise HTTPException(400, f"Only .egp files allowed (got {ext!r})")

    target_dir = _SAS_SESSIONS_ROOT / session_id
    target_dir.mkdir(parents=True, exist_ok=True)

    blocks_metadata = _extract_egp(file, target_dir)
    _write_manifest(target_dir, file.filename or "project.egp", blocks_metadata)
    return {
        "session_id": session_id,
        "source_file": file.filename,
        "blocks_extracted": len(blocks_metadata),
        "blocks_metadata": blocks_metadata,
        "folder": str(target_dir),
    }


@router.delete("/sas/{session_id}/{name}")
def delete_sas(session_id: str, name: str) -> dict:
    target = _SAS_SESSIONS_ROOT / session_id / name
    if not target.exists():
        raise HTTPException(404, f"{target} not found")
    target.unlink()
    return {"deleted": str(target)}


# ── /agent/docs/reindex ──────────────────────────────────────────────────────


@router.post("/docs/reindex")
def reindex_docs() -> dict:
    from src.agent.docs_index import DocsIndex, reset_default_index

    idx = DocsIndex().build()
    idx.save(_INDEX_PATH)
    reset_default_index()
    return {
        "sections": idx.section_count(),
        "docs_root": str(idx.docs_root),
        "index_path": str(_INDEX_PATH),
    }


@router.post("/docs/upload")
async def upload_docs(
    files: list[UploadFile] = File(...),
    subfolder: str = Form("uploads"),
) -> dict:
    safe = Path(subfolder).name or "uploads"
    target_dir = _DOCS_ROOT / safe
    target_dir.mkdir(parents=True, exist_ok=True)
    saved: list[str] = []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext != ".md":
            raise HTTPException(400, f"Only .md files allowed (got {ext!r})")
        target = target_dir / Path(f.filename or "").name
        with target.open("wb") as out:
            shutil.copyfileobj(f.file, out)
        saved.append(target.name)
    # Auto-rebuild
    from src.agent.docs_index import DocsIndex, reset_default_index
    idx = DocsIndex().build()
    idx.save(_INDEX_PATH)
    reset_default_index()
    return {
        "saved": saved,
        "subfolder": str(target_dir),
        "sections_after_reindex": idx.section_count(),
    }


# ── /agent/status ────────────────────────────────────────────────────────────


def _count_sas_sessions() -> dict:
    """Count SAS files per session."""
    sessions: dict[str, dict] = {}
    if _SAS_SESSIONS_ROOT.exists():
        for d in sorted(_SAS_SESSIONS_ROOT.iterdir()):
            if d.is_dir():
                files = sorted(str(p.name) for p in d.rglob("*.sas"))
                sessions[d.name] = {"count": len(files), "files": files}
    return sessions


@router.get("/status")
def status() -> dict:
    from src.agent.docs_index import get_default_index
    from src.agent.tools import TOOL_REGISTRY
    from src.knowledge import get_client

    client = get_client()
    backend = client.detect_backend()
    model = (
        client.litert_model if backend == "litert"
        else client.ollama_model if backend == "ollama"
        else "stub"
    )
    idx = get_default_index()
    return {
        "llm": {"backend": backend, "model": model},
        "tools": list(TOOL_REGISTRY.keys()),
        "sas_sessions": _count_sas_sessions(),
        "docs": {
            "root": str(_DOCS_ROOT),
            "sections": idx.section_count(),
            "index_path": str(_INDEX_PATH),
            "index_present": _INDEX_PATH.exists(),
        },
    }
