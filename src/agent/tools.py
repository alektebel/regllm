"""Tool registry for the SAS-diff agent.

Each tool is described by a JSON schema (the ``OpenAI / Ollama tools``
format) and a Python callable. The registry is the *single source of
truth* for what the LLM can do.

Tools are deliberately small and side-effect free — they read the
filesystem, never write — so the agent loop is safe to run on any user
question without further sandboxing.
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

logger = logging.getLogger(__name__)


# ── Filesystem layout ────────────────────────────────────────────────────────

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_SAS_ROOT = _PROJECT_ROOT / "data" / "sas"
_SAMPLES = _PROJECT_ROOT / "data" / "samples"
_DOCS_ROOT = _PROJECT_ROOT / "data" / "docs"
_CSV_V3 = _SAMPLES / "cycles_v3.csv"
_PK = "CICLO_ID"

_NUMERIC_COLS = {
    "PD_ESTIMADA", "LGD_ESTIMADA", "EAD", "DPDS", "STAGE_IFRS9",
    "PROVISION_PERIOD_MONTHS", "VENTANA_OBSERVACION_YEARS",
    "VENTANA_CALIBRACION_YEARS", "RATING_GRADO", "ECL",
    "LGD_FLOOR_APLICADO", "MOC", "CURE_FLAG", "RWA",
    "OR_EAD_TIT", "OR_EAD",  # potential demo fields
}


# ── Helpers ──────────────────────────────────────────────────────────────────


_SAS_SESSIONS_ROOT = _SAS_ROOT / "sessions"


def _load_sas(session_id: str = "default") -> str:
    """Load all `.sas` files for a session, falling back to bundled sample."""
    folder = _SAS_SESSIONS_ROOT / session_id
    if folder.exists():
        files = sorted(f for f in folder.rglob("*.sas") if f.name != "_project_manifest.json")
        if files:
            return "\n\n".join(f.read_text(encoding="utf-8") for f in files)
    sample = _SAMPLES / "sample_lgd.sas"
    if sample.exists():
        return sample.read_text(encoding="utf-8")
    return ""


def _coerce_row(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in row.items():
        ku = k.upper() if isinstance(k, str) else k
        if v in ("", None):
            continue
        if ku in _NUMERIC_COLS and isinstance(v, str):
            try:
                out[ku] = float(v)
                continue
            except ValueError:
                pass
        out[ku] = v
    return out


def _read_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return [_coerce_row(r) for r in csv.DictReader(f)]


def _truncate(obj: Any, limit: int = 4000) -> Any:
    """Compact a result so it doesn't blow up the LLM context."""
    s = json.dumps(obj, default=str)
    if len(s) <= limit:
        return obj
    return {"truncated": True, "preview": s[:limit] + "..."}


# ── Tool implementations ────────────────────────────────────────────────────


def _t_find_row(pk: str) -> dict[str, Any]:
    csv_path = _CSV_V3
    rows = _read_csv(csv_path)
    pk_u = pk.upper()
    for r in rows:
        if str(r.get(_PK, "")).upper() == pk_u:
            return {"found": True, "pk": pk, "row": r}
    return {"found": False, "pk": pk, "row": None}


def _approx_eq(a: Any, b: Any, tolerance: float) -> bool:
    if a == b:
        return True
    try:
        return abs(float(a) - float(b)) <= tolerance
    except (TypeError, ValueError):
        return False


def _t_find_rows_by_field_value(
    field: str,
    value: Any,
    tolerance: float = 1e-6,
    limit: int = 10,
) -> dict[str, Any]:
    csv_path = _CSV_V3
    rows = _read_csv(csv_path)
    field_u = field.upper()
    hits: list[dict[str, Any]] = []
    for r in rows:
        if _approx_eq(r.get(field_u), value, tolerance):
            hits.append({"pk": r.get(_PK), field_u: r.get(field_u)})
            if len(hits) >= limit:
                break
    return {"field": field_u, "value": value, "matches": hits, "total_scanned": len(rows)}


def _t_inspect_lineage(target: str, session_id: str = "default") -> dict[str, Any]:
    from src.sas_logic_tree import SASLogicTree
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    trace = tree.trace_lineage(nodes, target)
    return {
        "target": trace["target"],
        "session_id": session_id,
        "found": trace["found"],
        "ancestor_count": trace["ancestor_count"],
        "ancestors": trace["ancestors"][:80],
        "direct_predecessors": trace["direct_predecessors"][:40],
        "layers": trace["layers"][:12],
    }


def _t_trace_dependencies(
    target: str, session_id: str = "default", max_depth: int | None = None,
) -> dict[str, Any]:
    """Full BFS dependency trace with edges (expressions, data steps)."""
    from src.sas_logic_tree import SASLogicTree
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    trace = tree.trace_lineage(nodes, target, max_depth=max_depth)
    return {
        "target": trace["target"],
        "session_id": session_id,
        "found": trace["found"],
        "ancestor_count": trace["ancestor_count"],
        "layers": trace["layers"][:20],
        "edges": trace["edges"][:100],
        "depth": dict(list(trace["depth"].items())[:80]),
    }


def _t_get_sas_formula(field: str, session_id: str = "default") -> dict[str, Any]:
    """Return verified assignment expressions for a field from parsed SAS."""
    from src.sas_logic_tree import SASLogicTree, _vars_in_expr
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    trace = tree.trace_lineage(nodes, field)
    field_u = field.upper()

    definitions: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for edge in trace.get("edges", []):
        if edge.get("target", "").upper() != field_u:
            continue
        if edge.get("kind") != "assigns":
            continue
        expr = (edge.get("expr") or "").strip()
        if not expr:
            continue
        step = edge.get("data_step") or ""
        key = (step, expr)
        if key in seen:
            continue
        seen.add(key)
        definitions.append({
            "data_step": step,
            "expr": expr,
            "inputs": sorted(_vars_in_expr(expr)),
        })

    return {
        "field": field_u,
        "session_id": session_id,
        "found": bool(definitions) or trace.get("found", False),
        "definitions": definitions,
        "hint": (
            "Copia expr literalmente al responder. "
            "Si definitions está vacío, el campo no se asigna en esta sesión "
            "(puede venir de tabla de entrada)."
        ),
    }


def _t_trace_causal_chain(
    target: str,
    session_id: str = "default",
    max_depth: int | None = None,
    include_data: bool = True,
    data_limit: int = 5,
) -> dict[str, Any]:
    """Full upstream chain with formulas, SAS comments and optional DB samples."""
    from src.agent.causal_chain import build_causal_chain
    from src.agent.pipeline_store import sample_for_causal_chain
    from src.sas_logic_tree import SASLogicTree

    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    trace = tree.trace_lineage(nodes, target, max_depth=max_depth)
    out = build_causal_chain(trace, session_id, include_propagation=False)

    if include_data:
        chain_fields = [s["field"] for s in out.get("steps", [])]
        # MoC in CSV vs MOC in compiler
        chain_fields = [
            "MoC" if f == "MOC" else f for f in chain_fields
        ]
        data = sample_for_causal_chain(
            session_id, target, chain_fields, limit=data_limit,
        )
        out["data"] = data
        if data.get("rows_missing_target") or data.get("rows_with_target"):
            out["narrative"] += "\n\nValores reales en BD (muestra):"
            for label, key in (
                ("missing", "rows_missing_target"),
                ("con valor", "rows_with_target"),
            ):
                for row in data.get(key, [])[:3]:
                    pk = row.get(_PK, "?")
                    vals = ", ".join(
                        f"{k}={json.dumps(v) if v is not None else 'null'}"
                        for k, v in row.items() if k != _PK
                    )
                    out["narrative"] += f"\n  • {pk} ({label}): {vals}"

    return out


def _t_query_cycle_data(
    session_id: str = "default",
    ciclo_id: str | None = None,
    fields: list[str] | None = None,
    missing_field: str | None = None,
    filter_field: str | None = None,
    filter_value: str | float | None = None,
    limit: int = 10,
) -> dict[str, Any]:
    """Query actual pipeline output rows for the session (ground-truth values)."""
    from src.agent.pipeline_store import query_cycle_data
    return query_cycle_data(
        session_id,
        ciclo_id=ciclo_id,
        fields=fields,
        missing_field=missing_field,
        filter_field=filter_field,
        filter_value=filter_value,
        limit=limit,
    )


def _t_compute_shapley(
    pk: str, target: str, session_id: str = "default",
) -> dict[str, Any]:
    """Shapley value attribution for a specific cycle and target field."""
    from src.sas_diff import explain_field_diff
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    rows = _read_csv(_CSV_V3)
    pk_u = pk.upper()
    row = next((r for r in rows if str(r.get(_PK, "")).upper() == pk_u), None)
    if not row:
        return {"error": f"row {pk} not found in sample CSV"}
    rep = explain_field_diff(
        sas_code=sas, row_v2=row, row_v3=row,
        target=target, method="shapley",
    )
    out = rep.to_dict()
    return {
        "target": out["target"],
        "session_id": session_id,
        "suspects": out["suspects"][:8],
        "field_deltas": out["field_deltas"][:20],
        "shapley_top": (out.get("shapley", {}) or {}).get("attributions", [])[:8],
    }


def _t_search_docs(query: str, k: int = 5) -> dict[str, Any]:
    from src.agent.docs_index import get_default_index
    idx = get_default_index()
    hits = idx.search(query, k=k)
    return {
        "query": query,
        "k": k,
        "hits": [h.to_dict() for h in hits],
        "doc_count": idx.section_count(),
    }


def _t_get_field_definition(field: str) -> dict[str, Any]:
    from src.agent.docs_index import get_default_index
    idx = get_default_index()
    res = idx.get_field_definition(field)
    return {"field": field.upper(), "result": res}


def _t_search_regulation(query: str, k: int = 3) -> dict[str, Any]:
    try:
        from src.knowledge import GraphRAG, collect_evidence, field_subgraph, linearise_subgraph
        rag = GraphRAG(
            graph_path=Path("data/regulation/graph.json"),
            changelog_dir=Path("data/regulation"),
            schema_files=[],
        )
        g = rag.graph
        q_u = query.upper().strip()
        sub = field_subgraph(g, q_u, hops=2)
        return {
            "query": query,
            "evidence": collect_evidence(g, q_u),
            "linearised": linearise_subgraph(sub, q_u)[:1500],
        }
    except Exception as e:
        logger.warning("regulation tool failed: %s", e)
        return {"error": str(e), "query": query}


def _t_search_changelog(query: str, k: int = 3) -> dict[str, Any]:
    try:
        from src.knowledge import GraphRAG, collect_evidence, field_subgraph, linearise_subgraph
        rag = GraphRAG()
        g = rag.graph
        # Treat the query as a field name first; fall back to a substring scan over labels.
        q_u = query.upper().strip()
        sub = field_subgraph(g, q_u, hops=2)
        out = {
            "query": query,
            "evidence": collect_evidence(g, q_u),
            "linearised": linearise_subgraph(sub, q_u)[:1500],
        }
        return out
    except Exception as e:
        logger.warning("changelog tool failed: %s", e)
        return {"error": str(e), "query": query}


def _t_enrich_fields(
    fields: list[str],
    include_docs: bool = True,
    include_regulation: bool = False,
) -> dict[str, Any]:
    """Batch-enrich a list of field names with doc definitions and regulation."""
    results: dict[str, Any] = {}
    for field in fields[:15]:
        field_u = field.upper().strip()
        entry: dict[str, Any] = {"field": field_u}
        if include_docs:
            entry["definition"] = _t_get_field_definition(field_u).get("result")
        if include_regulation:
            entry["regulation"] = _t_search_regulation(field_u)
        results[field_u] = entry
    return {"enriched": results, "count": len(results)}


def _t_get_schema_context(
    query: str, tables: list[str] | None = None,
) -> dict[str, Any]:
    """Extract relevant table DDL from the reference schema."""
    import re as _re
    schema_path = _SAMPLES / "irb_schema.sql"
    if not schema_path.exists():
        return {"error": "irb_schema.sql not found"}
    schema_text = schema_path.read_text(encoding="utf-8")
    table_blocks = _re.split(r"(?=CREATE TABLE)", schema_text)
    results = []
    query_terms = query.upper().split()
    for block in table_blocks:
        if not block.strip():
            continue
        m = _re.match(r"CREATE TABLE\s+(\w+)", block)
        if not m:
            continue
        tname = m.group(1)
        if tables and tname.lower() not in [t.lower() for t in tables]:
            continue
        if tables or any(term in block.upper() for term in query_terms):
            results.append({"table": tname, "ddl": block[:2000]})
    return {"query": query, "tables": results[:10]}


def _t_describe_egp_project(session_id: str = "default") -> dict[str, Any]:
    """Read the project manifest and return structure with code previews."""
    manifest_path = _SAS_SESSIONS_ROOT / session_id / "_project_manifest.json"
    if not manifest_path.exists():
        # Fall back to listing .sas files in the session folder
        folder = _SAS_SESSIONS_ROOT / session_id
        if not folder.exists():
            return {"error": f"no SAS session found: {session_id}"}
        files = sorted(f for f in folder.rglob("*.sas"))
        blocks = []
        for i, f in enumerate(files):
            code = f.read_text(encoding="utf-8")
            blocks.append({
                "filename": f.name,
                "task_name": f.stem,
                "order": i,
                "line_count": len(code.splitlines()),
                "preview": "\n".join(code.splitlines()[:10]),
            })
        return {"session_id": session_id, "source": "file_listing", "blocks": blocks}

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    blocks = []
    for b in manifest.get("blocks", []):
        sas_path = _SAS_SESSIONS_ROOT / session_id / b["filename"]
        preview = ""
        if sas_path.exists():
            preview = "\n".join(sas_path.read_text(encoding="utf-8").splitlines()[:10])
        blocks.append({
            **b,
            "preview": preview,
        })
    return {
        "session_id": session_id,
        "source_file": manifest.get("source_file", ""),
        "extracted_at": manifest.get("extracted_at", ""),
        "block_count": len(blocks),
        "blocks": blocks,
    }


def _t_validate_sas(session_id: str = "default") -> dict[str, Any]:
    """Run static validation on parsed SAS code."""
    from src.sas_logic_tree import SASLogicTree
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    diags = tree.validate(nodes)
    return {"session_id": session_id, "diagnostics": diags[:50], "total": len(diags)}


# ── Registry ────────────────────────────────────────────────────────────────


@dataclass
class ToolSpec:
    name: str
    description: str
    parameters: dict[str, Any]
    fn: Callable[..., Any]

    def schema(self) -> dict[str, Any]:
        """Return the OpenAI/Ollama-format tool schema."""
        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": self.parameters,
            },
        }


TOOL_REGISTRY: dict[str, ToolSpec] = {
    "find_row": ToolSpec(
        name="find_row",
        description=(
            "Fetch a specific cycle's full row from sample data. "
            "Use this once you know the primary key (CICLO_ID) to inspect "
            "every input field for that cycle."
        ),
        parameters={
            "type": "object",
            "properties": {
                "pk": {"type": "string", "description": "Cycle primary key, e.g. 'CIC_00031'"},
            },
            "required": ["pk"],
        },
        fn=_t_find_row,
    ),
    "find_rows_by_field_value": ToolSpec(
        name="find_rows_by_field_value",
        description=(
            "Find cycles where a given field has approximately a given value. "
            "Useful when the user mentions a value but no PK."
        ),
        parameters={
            "type": "object",
            "properties": {
                "field": {"type": "string", "description": "Field name, e.g. 'OR_EAD_TIT'"},
                "value": {
                    "description": "Numeric or string value to match.",
                    "anyOf": [{"type": "number"}, {"type": "string"}, {"type": "boolean"}],
                },
                "tolerance": {"type": "number", "default": 1e-6, "description": "Numeric tolerance"},
                "limit": {"type": "integer", "default": 10},
            },
            "required": ["field", "value"],
        },
        fn=_t_find_rows_by_field_value,
    ),
    "inspect_lineage": ToolSpec(
        name="inspect_lineage",
        description=(
            "Return the data-flow ancestors of a target field in the SAS "
            "pipeline. Use this to understand which input fields *can* affect "
            "the target before computing attributions."
        ),
        parameters={
            "type": "object",
            "properties": {
                "target": {"type": "string"},
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
            },
            "required": ["target"],
        },
        fn=_t_inspect_lineage,
    ),
    "trace_dependencies": ToolSpec(
        name="trace_dependencies",
        description=(
            "Trace the full dependency chain of a target field backwards through "
            "the SAS pipeline using BFS. Returns all ancestor fields grouped by "
            "hop distance, with edges showing HOW each dependency flows "
            "(assignment expression, data step, edge kind)."
        ),
        parameters={
            "type": "object",
            "properties": {
                "target": {"type": "string", "description": "Target field to trace backwards from."},
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
                "max_depth": {"type": "integer", "description": "Max BFS depth (omit for unlimited)."},
            },
            "required": ["target"],
        },
        fn=_t_trace_dependencies,
    ),
    "get_sas_formula": ToolSpec(
        name="get_sas_formula",
        description=(
            "Return the exact SAS assignment expression(s) for a field in the "
            "current session pipeline. Use this BEFORE stating any formula "
            "(MoC, ECL, LGD, EAD). Do not use get_field_definition for "
            "pipeline formulas — that tool reads markdown docs, not SAS code."
        ),
        parameters={
            "type": "object",
            "properties": {
                "field": {"type": "string", "description": "Field name, e.g. 'MoC' or 'ECL'."},
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
            },
            "required": ["field"],
        },
        fn=_t_get_sas_formula,
    ),
    "trace_causal_chain": ToolSpec(
        name="trace_causal_chain",
        description=(
            "Full causal chain for root-cause analysis: upstream fields with "
            "assignment expr, SAS comments (BUG/missing), and actual DB row "
            "samples in `data` (include_data=true by default). Use for "
            "'missing', 'investiga', 'causa raíz'."
        ),
        parameters={
            "type": "object",
            "properties": {
                "target": {"type": "string", "description": "Symptom field, e.g. 'ECL' or 'MoC'."},
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
                "max_depth": {"type": "integer", "description": "Max BFS depth (omit for unlimited)."},
                "include_data": {"type": "boolean", "default": True, "description": "Attach actual pipeline output rows."},
                "data_limit": {"type": "integer", "default": 5, "description": "Max sample rows per category."},
            },
            "required": ["target"],
        },
        fn=_t_trace_causal_chain,
    ),
    "query_cycle_data": ToolSpec(
        name="query_cycle_data",
        description=(
            "Query actual values from the pipeline output table for this session. "
            "Use to inspect specific cycles (ciclo_id), find rows where a field "
            "is missing (missing_field), or filter by field=value. Returns real "
            "DB values — not inferred propagation."
        ),
        parameters={
            "type": "object",
            "properties": {
                "session_id": {"type": "string", "default": "default"},
                "ciclo_id": {"type": "string", "description": "Primary key, e.g. CIC_00042."},
                "fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Columns to return (default: all).",
                },
                "missing_field": {"type": "string", "description": "Only rows where this field is null/missing."},
                "filter_field": {"type": "string"},
                "filter_value": {"description": "Value to match; use 'missing' for null."},
                "limit": {"type": "integer", "default": 10},
            },
        },
        fn=_t_query_cycle_data,
    ),
    "compute_shapley": ToolSpec(
        name="compute_shapley",
        description=(
            "Run Shapley value attribution for a specific cycle and target "
            "field. Returns ranked suspect fields and top contributions."
        ),
        parameters={
            "type": "object",
            "properties": {
                "pk": {"type": "string"},
                "target": {"type": "string"},
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
            },
            "required": ["pk", "target"],
        },
        fn=_t_compute_shapley,
    ),
    "search_docs": ToolSpec(
        name="search_docs",
        description=(
            "Full-text search over the markdown corpus under data/docs/ "
            "(table dictionaries, flux explanations, field semantics)."
        ),
        parameters={
            "type": "object",
            "properties": {
                "query": {"type": "string"},
                "k": {"type": "integer", "default": 5},
            },
            "required": ["query"],
        },
        fn=_t_search_docs,
    ),
    "get_field_definition": ToolSpec(
        name="get_field_definition",
        description=(
            "Look up the semantic definition of a specific field in the "
            "markdown docs corpus."
        ),
        parameters={
            "type": "object",
            "properties": {"field": {"type": "string"}},
            "required": ["field"],
        },
        fn=_t_get_field_definition,
    ),
    "search_changelog": ToolSpec(
        name="search_changelog",
        description=(
            "Search the database change-log GraphRAG for documented V2→V3 "
            "changes that mention a field. Returns evidence sections plus a "
            "linearised subgraph."
        ),
        parameters={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Field name preferred; otherwise free text."},
                "k": {"type": "integer", "default": 3},
            },
            "required": ["query"],
        },
        fn=_t_search_changelog,
    ),
    "search_regulation": ToolSpec(
        name="search_regulation",
        description=(
            "Search the official regulation GraphRAG (Circular 6/2016 BdE) "
            "for articles governing provision periods on recovery cycles. "
            "Use this to ground answers in regulatory citations — especially "
            "for questions about PROVISION_PERIOD_MONTHS, LGD_FLOOR_APLICADO, "
            "STAGE_IFRS9 requirements, or CURE_FLAG conditions."
        ),
        parameters={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Field name or regulatory concept, e.g. 'PROVISION_PERIOD_MONTHS' or 'liberacion provisiones'."},
                "k": {"type": "integer", "default": 3},
            },
            "required": ["query"],
        },
        fn=_t_search_regulation,
    ),
    "enrich_fields": ToolSpec(
        name="enrich_fields",
        description=(
            "Batch-enrich a list of field names with documentation definitions "
            "and optionally regulatory context. Use after trace_dependencies "
            "to understand the semantic meaning of ancestor fields."
        ),
        parameters={
            "type": "object",
            "properties": {
                "fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of field names to enrich.",
                },
                "include_docs": {"type": "boolean", "default": True},
                "include_regulation": {"type": "boolean", "default": False},
            },
            "required": ["fields"],
        },
        fn=_t_enrich_fields,
    ),
    "get_schema_context": ToolSpec(
        name="get_schema_context",
        description=(
            "Extract relevant table definitions from the IRB reference schema "
            "(irb_schema.sql). Use this before generating SAS/PROC SQL code to "
            "know the correct table and column names."
        ),
        parameters={
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Free-text query or field name."},
                "tables": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional: restrict to specific table names.",
                },
            },
            "required": ["query"],
        },
        fn=_t_get_schema_context,
    ),
    "validate_sas": ToolSpec(
        name="validate_sas",
        description=(
            "Run static analysis on the parsed SAS code to detect potential "
            "errors: undefined variables, missing datasets, type mismatches, "
            "and schema violations against the IRB reference schema."
        ),
        parameters={
            "type": "object",
            "properties": {
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
            },
        },
        fn=_t_validate_sas,
    ),
    "describe_egp_project": ToolSpec(
        name="describe_egp_project",
        description=(
            "Describe the structure of an uploaded .egp (SAS Enterprise Guide) "
            "project. Returns task names, ordering, block types, line counts, "
            "and a code preview of each block. Call this first after an .egp upload."
        ),
        parameters={
            "type": "object",
            "properties": {
                "session_id": {"type": "string", "default": "default", "description": "Session with uploaded .egp."},
            },
        },
        fn=_t_describe_egp_project,
    ),
}


# ── Knowledge Graph tools ─────────────────────────────────────────────────────

_graph_store = None


def _get_graph_store():
    """Lazy-load the GraphStore singleton."""
    global _graph_store
    if _graph_store is None:
        try:
            from src.knowledge.graph_store import GraphStore
            _graph_store = GraphStore()
        except Exception as e:
            logger.warning("GraphStore unavailable: %s", e)
    return _graph_store


def _t_query_regulation(
    concept: str = "",
    article: str = "",
    hops: int = 2,
) -> dict[str, Any]:
    """Multi-hop traversal from a regulation concept or article."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import NodeType, EdgeType

    # Find the starting node(s)
    nodes = []
    if article:
        nodes = store.search_nodes(node_types=[NodeType.REGULATION], label_contains=article, limit=5)
    if not nodes and concept:
        nodes = store.search_nodes(label_contains=concept, limit=5)
    if not nodes:
        return {"found": False, "query": concept or article, "results": []}

    # Get subgraph around the first match
    center = nodes[0]
    sg = store.subgraph(center.id, hops=hops)
    return {
        "found": True,
        "center": center.model_dump(),
        "subgraph": sg,
    }


def _t_find_governing_rules(
    column: str,
    table: str | None = None,
) -> dict[str, Any]:
    """Find all regulations that govern a database column."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import EdgeType, NodeType

    col_id = f"col:{column.upper()}"
    node = store.get_node(col_id)
    if node is None:
        # Try concept
        col_id = f"concept:{column.upper()}"
        node = store.get_node(col_id)
    if node is None:
        return {"found": False, "column": column, "rules": []}

    # Get inbound edges (regulations that GOVERN this column)
    edges = store.get_edges(col_id, direction="in")
    rules = []
    for e in edges:
        if e.edge_type in (EdgeType.GOVERNS, EdgeType.VALIDATES, EdgeType.IMPLEMENTS):
            src_node = store.get_node(e.src_id)
            if src_node:
                rules.append({
                    "regulation": src_node.model_dump(),
                    "edge_type": e.edge_type.value,
                    "confidence": e.confidence,
                })

    # Also check for exceptions
    exception_edges = store.get_edges(col_id, edge_types=[EdgeType.EXCEPTION_TO], direction="both")
    exceptions = []
    for e in exception_edges:
        other_id = e.dst_id if e.src_id == col_id else e.src_id
        other = store.get_node(other_id)
        if other:
            exceptions.append(other.model_dump())

    return {
        "found": True,
        "column": column,
        "rules": rules,
        "exceptions": exceptions,
    }


def _t_trace_regulation_chain(
    article_id: str,
    target_column: str,
) -> dict[str, Any]:
    """Shortest path from a regulation article to a database column."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import NodeType

    # Find regulation node
    reg_nodes = store.search_nodes(node_types=[NodeType.REGULATION], label_contains=article_id, limit=3)
    if not reg_nodes:
        return {"found": False, "error": f"Regulation '{article_id}' not found"}

    # Find column node
    col_id = f"col:{target_column.upper()}"
    col_node = store.get_node(col_id)
    if col_node is None:
        return {"found": False, "error": f"Column '{target_column}' not found"}

    path = store.shortest_path(reg_nodes[0].id, col_id, max_hops=5)
    if path is None:
        return {"found": False, "from": reg_nodes[0].id, "to": col_id, "path": None}

    return {
        "found": True,
        "from": reg_nodes[0].id,
        "to": col_id,
        "path": path,
        "hop_count": len([p for p in path if p["kind"] == "edge"]),
    }


def _t_search_experience(
    query: str,
    k: int = 5,
) -> dict[str, Any]:
    """Search the experience/insight KB for past validation learnings."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import NodeType

    # Search by label match in experience and insight nodes
    results = store.search_nodes(
        node_types=[NodeType.EXPERIENCE, NodeType.INSIGHT, NodeType.CONCEPT],
        label_contains=query,
        limit=k * 2,  # fetch extra, then sort by priority
    )
    # Sort insights by priority (feedback > auto-discovered), experiences last
    def _sort_key(n: Any) -> tuple[float, str]:
        priority = getattr(n, "priority", 0.5) if hasattr(n, "priority") else 0.5
        return (-priority, n.label)
    results.sort(key=_sort_key)
    results = results[:k]
    return {
        "query": query,
        "results": [n.model_dump() for n in results],
        "count": len(results),
    }


def _t_save_insight(
    insight: str,
    related_fields: list[str] | None = None,
    related_articles: list[str] | None = None,
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """Save a new insight from the current validation session."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import Insight, KGEdge, EdgeType
    import hashlib
    from datetime import datetime

    # Generate a unique ID
    ts = datetime.now().isoformat()
    hash_input = f"{insight}:{ts}"
    insight_id = f"insight:{hashlib.sha256(hash_input.encode()).hexdigest()[:12]}"

    store.add_node(Insight(
        id=insight_id,
        label=insight[:100],
        summary=insight,
        tags=tags or [],
    ))

    # Link to related columns
    edges_created = 0
    for field in (related_fields or []):
        col_id = f"col:{field.upper()}"
        concept_id = f"concept:{field.upper()}"
        # Link to whichever exists
        target = col_id if store.get_node(col_id) else concept_id
        if store.get_node(target):
            store.add_edge(KGEdge(
                src_id=insight_id,
                dst_id=target,
                edge_type=EdgeType.RELATES_TO,
            ))
            edges_created += 1

    # Link to related regulations
    for article in (related_articles or []):
        from src.knowledge.ontology import NodeType
        reg_nodes = store.search_nodes(node_types=[NodeType.REGULATION], label_contains=article, limit=1)
        if reg_nodes:
            store.add_edge(KGEdge(
                src_id=insight_id,
                dst_id=reg_nodes[0].id,
                edge_type=EdgeType.RELATES_TO,
            ))
            edges_created += 1

    return {
        "saved": True,
        "id": insight_id,
        "edges_created": edges_created,
    }


def _t_save_feedback(
    feedback_type: str,
    content: str,
    original_claim: str = "",
    corrected_understanding: str = "",
    related_fields: list[str] | None = None,
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """Save user feedback as a high-priority Insight node."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import Insight, KGEdge, EdgeType
    import hashlib
    from datetime import datetime

    ts = datetime.now().isoformat()
    hash_input = f"feedback:{content}:{ts}"
    fb_id = f"insight:{hashlib.sha256(hash_input.encode()).hexdigest()[:12]}"

    store.add_node(Insight(
        id=fb_id,
        label=f"[feedback] {content[:90]}",
        summary=content,
        tags=["feedback"] + (tags or []),
        priority=1.0,
        feedback_type=feedback_type,
        original_claim=original_claim,
        corrected_understanding=corrected_understanding,
    ))

    edges_created = 0
    for field in (related_fields or []):
        col_id = f"col:{field.upper()}"
        concept_id = f"concept:{field.upper()}"
        target = col_id if store.get_node(col_id) else concept_id
        if store.get_node(target):
            store.add_edge(KGEdge(
                src_id=fb_id,
                dst_id=target,
                edge_type=EdgeType.RELATES_TO,
            ))
            edges_created += 1

    return {"saved": True, "id": fb_id, "edges_created": edges_created}


def _t_backtrace_sas_field(
    target: str,
    session_id: str = "default",
) -> dict[str, Any]:
    """Trace a field backward through SAS logic, group leaf ancestors by source table."""
    from src.sas_logic_tree import SASLogicTree
    sas = _load_sas(session_id)
    if not sas:
        return {"error": f"no SAS found for session {session_id}"}
    tree = SASLogicTree()
    nodes = tree.parse(sas)
    trace = tree.trace_lineage(nodes, target)
    if not trace["found"]:
        return {"target": target, "found": False, "input_tables": []}

    # Group leaf ancestors (those with no further predecessors) by data step
    ancestors = trace.get("ancestors", [])
    layers = trace.get("layers", [])
    edges = trace.get("edges", [])

    # Find leaf fields (appear in deepest layers or have no outgoing edges in trace)
    all_sources: set[str] = set()
    for a in ancestors:
        all_sources.add(a.upper() if isinstance(a, str) else str(a).upper())

    # Group by data step context from edges
    table_fields: dict[str, set[str]] = {}
    for edge in edges:
        src = edge.get("src", edge.get("from", ""))
        step = edge.get("data_step", "unknown")
        if isinstance(src, str):
            table_fields.setdefault(step, set()).add(src.upper())

    # Build input_tables list
    input_tables = []
    for step, fields in sorted(table_fields.items()):
        input_tables.append({
            "table": step,
            "fields": sorted(fields),
            "field_count": len(fields),
        })

    return {
        "target": target,
        "session_id": session_id,
        "found": True,
        "ancestor_count": trace["ancestor_count"],
        "input_tables": input_tables[:20],
        "layers": layers[:12],
        "direct_predecessors": trace["direct_predecessors"][:30],
    }


def _t_create_investigation_plan(
    target_field: str,
    problem: str,
    steps: list[str],
    related_fields: list[str] | None = None,
    cycles: list[str] | None = None,
) -> dict[str, Any]:
    """Persist an investigation plan as an Experience node in the KG."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.ontology import Experience, KGEdge, EdgeType
    import hashlib
    from datetime import datetime

    ts = datetime.now().isoformat()
    plan_id = f"exp:plan_{hashlib.sha256(f'{problem}:{ts}'.encode()).hexdigest()[:12]}"

    plan_text = f"Investigation: {problem}\nTarget: {target_field}\n"
    if cycles:
        plan_text += f"Cycles: {', '.join(cycles)}\n"
    plan_text += "Steps:\n" + "\n".join(f"  {i+1}. {s}" for i, s in enumerate(steps))

    store.add_node(Experience(
        id=plan_id,
        label=f"[plan] {problem[:80]}",
        session_id="",
        question=problem,
        answer_summary=plan_text,
        timestamp=ts,
    ))

    # Link to related fields
    edges_created = 0
    for field in (related_fields or []) + [target_field]:
        for prefix in ("col:", "concept:"):
            target_id = f"{prefix}{field.upper()}"
            if store.get_node(target_id):
                store.add_edge(KGEdge(
                    src_id=plan_id,
                    dst_id=target_id,
                    edge_type=EdgeType.RELATES_TO,
                ))
                edges_created += 1
                break

    return {
        "saved": True,
        "id": plan_id,
        "plan": plan_text,
        "edges_created": edges_created,
    }


def _t_merge_nodes(source: str, target: str) -> dict[str, Any]:
    """Merge duplicate insight nodes (source into target)."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import merge_nodes

    return merge_nodes(store, source, target)


def _t_add_tags(node: str, tags: list[str]) -> dict[str, Any]:
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import add_tags

    return add_tags(store, node, tags)


def _t_link_nodes(source: str, target: str, relation: str = "RELATES_TO") -> dict[str, Any]:
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import link_nodes

    return link_nodes(store, source, target, relation=relation)


def _t_flag_conflict(node_a: str, node_b: str, reason: str = "") -> dict[str, Any]:
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import flag_conflict

    return flag_conflict(store, node_a, node_b, reason=reason)


def _t_delete_node(node: str) -> dict[str, Any]:
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    if store.get_node(node) is None:
        return {"deleted": False, "error": f"node not found: {node}"}
    store.delete_node(node)
    return {"deleted": True, "node": node}


def _t_create_concept(
    label: str,
    definition: str,
    covers: list[str],
    links_to: list[str] | None = None,
) -> dict[str, Any]:
    """Abstract related insights into a new Concept (reduces graph entropy)."""
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import create_concept

    return create_concept(store, label, definition, covers, links_to=links_to)


def _t_link_to_concept(insight_id: str, concept_id: str) -> dict[str, Any]:
    store = _get_graph_store()
    if store is None:
        return {"error": "Knowledge graph not available"}
    from src.knowledge.memory_ops import link_to_concept

    return link_to_concept(store, insight_id, concept_id)


def _t_formulate_data_request(
    extractions: list[dict[str, Any]],
    reason: str,
) -> dict[str, Any]:
    """Structure data requirements for the user to upload.

    This is a pass-through tool: it formats the request for the UI to render
    as a data upload card. Each extraction specifies a table, fields, optional
    filters, cycles, and a human description.
    """
    formatted = []
    for ext in extractions[:10]:
        formatted.append({
            "table": ext.get("table", ""),
            "fields": ext.get("fields", []),
            "filters": ext.get("filters", ""),
            "cycles": ext.get("cycles", []),
            "description": ext.get("description", ""),
        })
    return {
        "data_request": True,
        "reason": reason,
        "extractions": formatted,
        "count": len(formatted),
    }


def _t_analyze_uploaded_data(
    file_path: str,
    target_fields: list[str] | None = None,
    cycle_filter: str | None = None,
    limit_rows: int = 50,
) -> dict[str, Any]:
    """Read a CSV/XLSX from data/uploads/ and return columns, stats, preview rows."""
    _UPLOADS = _PROJECT_ROOT / "data" / "uploads"
    # Resolve relative to uploads dir, prevent path traversal
    fp = (_UPLOADS / file_path).resolve()
    if not str(fp).startswith(str(_UPLOADS.resolve())):
        return {"error": "Invalid file path"}
    if not fp.exists():
        return {"error": f"File not found: {file_path}"}

    ext = fp.suffix.lower()
    try:
        if ext == ".csv":
            rows = _read_csv(fp)
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return {"error": "Empty workbook"}
            header = [str(c.value or f"col_{i}") for i, c in enumerate(next(ws.iter_rows(max_row=1)))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(header, row)))
            wb.close()
        else:
            return {"error": f"Unsupported format: {ext}"}
    except Exception as e:
        return {"error": f"Parse error: {e}"}

    columns = list(rows[0].keys()) if rows else []

    # Apply cycle filter
    if cycle_filter and rows:
        cycle_u = cycle_filter.upper()
        rows = [r for r in rows if any(str(v).upper() == cycle_u for v in r.values())]

    # Filter to target fields if specified
    if target_fields and rows:
        tf_set = {f.upper() for f in target_fields}
        # Keep PK columns + target fields
        keep = {c for c in columns if c.upper() in tf_set or "CICLO" in c.upper() or "ID" in c.upper()}
        rows = [{k: v for k, v in r.items() if k in keep} for r in rows]

    # Basic stats for numeric columns
    stats: dict[str, dict[str, Any]] = {}
    for col in columns[:30]:
        vals = [r.get(col) for r in rows if r.get(col) is not None]
        nums = []
        for v in vals:
            try:
                nums.append(float(v))
            except (TypeError, ValueError):
                pass
        if nums:
            stats[col] = {
                "count": len(nums),
                "min": min(nums),
                "max": max(nums),
                "mean": round(sum(nums) / len(nums), 6),
            }

    return {
        "file": file_path,
        "columns": columns,
        "row_count": len(rows),
        "stats": stats,
        "preview": rows[:limit_rows],
    }


TOOL_REGISTRY["query_regulation"] = ToolSpec(
    name="query_regulation",
    description=(
        "Multi-hop traversal of the regulatory knowledge graph. "
        "Given a concept name or article identifier, retrieves the "
        "subgraph neighbourhood (regulations, concepts, columns, "
        "validation rules) within the specified hop distance."
    ),
    parameters={
        "type": "object",
        "properties": {
            "concept": {"type": "string", "description": "Concept name or keyword, e.g. 'LGD floor'"},
            "article": {"type": "string", "description": "Article identifier, e.g. 'art15' or 'Artículo 15'"},
            "hops": {"type": "integer", "default": 2, "description": "Max traversal depth"},
        },
    },
    fn=_t_query_regulation,
)

TOOL_REGISTRY["find_governing_rules"] = ToolSpec(
    name="find_governing_rules",
    description=(
        "Find all regulation articles and validation rules that constrain "
        "a given database column. Returns governing rules with confidence "
        "scores and any known exceptions."
    ),
    parameters={
        "type": "object",
        "properties": {
            "column": {"type": "string", "description": "Column name, e.g. 'LGD_ESTIMADA'"},
            "table": {"type": "string", "description": "Optional table name to narrow the search."},
        },
        "required": ["column"],
    },
    fn=_t_find_governing_rules,
)

TOOL_REGISTRY["trace_regulation_chain"] = ToolSpec(
    name="trace_regulation_chain",
    description=(
        "Find the shortest path in the knowledge graph from a regulation "
        "article to a database column, showing all intermediate nodes "
        "(concepts, interpretations, validation rules). Use this to "
        "understand HOW a regulation constrains a specific column."
    ),
    parameters={
        "type": "object",
        "properties": {
            "article_id": {"type": "string", "description": "Regulation article, e.g. 'art15'"},
            "target_column": {"type": "string", "description": "Column name, e.g. 'LGD_ESTIMADA'"},
        },
        "required": ["article_id", "target_column"],
    },
    fn=_t_trace_regulation_chain,
)

TOOL_REGISTRY["backtrace_sas_field"] = ToolSpec(
    name="backtrace_sas_field",
    description=(
        "Trace a target field backward through SAS logic and group leaf "
        "ancestor fields by source table/data step. Use this to understand "
        "which input tables and fields feed into a target calculation, and "
        "to formulate data requirements for the user to upload."
    ),
    parameters={
        "type": "object",
        "properties": {
            "target": {"type": "string", "description": "Target field to trace backwards from."},
            "session_id": {"type": "string", "default": "default", "description": "Session with uploaded SAS code."},
        },
        "required": ["target"],
    },
    fn=_t_backtrace_sas_field,
)

TOOL_REGISTRY["create_investigation_plan"] = ToolSpec(
    name="create_investigation_plan",
    description=(
        "Persist an investigation plan for a validation case. Call this after "
        "you have gathered context from backtrace, regulation, and experience "
        "tools. The plan is saved as an Experience node in the KG so future "
        "sessions can recall it."
    ),
    parameters={
        "type": "object",
        "properties": {
            "target_field": {"type": "string", "description": "Main field under investigation."},
            "problem": {"type": "string", "description": "Problem description from the user."},
            "steps": {
                "type": "array", "items": {"type": "string"},
                "description": "Ordered investigation steps.",
            },
            "related_fields": {
                "type": "array", "items": {"type": "string"},
                "description": "Other fields involved.",
            },
            "cycles": {
                "type": "array", "items": {"type": "string"},
                "description": "Cycle IDs or periods mentioned.",
            },
        },
        "required": ["target_field", "problem", "steps"],
    },
    fn=_t_create_investigation_plan,
)

TOOL_REGISTRY["formulate_data_request"] = ToolSpec(
    name="formulate_data_request",
    description=(
        "Structure a data request for the user. After backtracing dependencies, "
        "call this to tell the user exactly which tables, fields, and cycles "
        "they need to extract and upload. The UI renders this as an upload card."
    ),
    parameters={
        "type": "object",
        "properties": {
            "extractions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "table": {"type": "string"},
                        "fields": {"type": "array", "items": {"type": "string"}},
                        "filters": {"type": "string"},
                        "cycles": {"type": "array", "items": {"type": "string"}},
                        "description": {"type": "string"},
                    },
                },
                "description": "List of data extractions needed.",
            },
            "reason": {"type": "string", "description": "Why this data is needed."},
        },
        "required": ["extractions", "reason"],
    },
    fn=_t_formulate_data_request,
)

TOOL_REGISTRY["analyze_uploaded_data"] = ToolSpec(
    name="analyze_uploaded_data",
    description=(
        "Read and analyze a CSV or XLSX file uploaded by the user. Returns "
        "column names, basic statistics, and a preview of rows. Optionally "
        "filter by cycle and target fields."
    ),
    parameters={
        "type": "object",
        "properties": {
            "file_path": {"type": "string", "description": "Relative path within data/uploads/"},
            "target_fields": {
                "type": "array", "items": {"type": "string"},
                "description": "Fields to focus on (keeps PK + these).",
            },
            "cycle_filter": {"type": "string", "description": "Filter rows to this cycle ID."},
            "limit_rows": {"type": "integer", "default": 50},
        },
        "required": ["file_path"],
    },
    fn=_t_analyze_uploaded_data,
)

TOOL_REGISTRY["save_feedback"] = ToolSpec(
    name="save_feedback",
    description=(
        "Save user feedback (correction, confirmation, or clarification) as a "
        "high-priority Insight node. Call this when the user corrects the agent "
        "or confirms a non-obvious finding. Feedback is recalled first in "
        "future search_experience results."
    ),
    parameters={
        "type": "object",
        "properties": {
            "feedback_type": {
                "type": "string",
                "enum": ["correction", "confirmation", "clarification"],
                "description": "Type of feedback.",
            },
            "content": {"type": "string", "description": "The feedback content."},
            "original_claim": {"type": "string", "description": "What was originally said/concluded."},
            "corrected_understanding": {"type": "string", "description": "The correct understanding."},
            "related_fields": {
                "type": "array", "items": {"type": "string"},
                "description": "Fields this feedback relates to.",
            },
            "tags": {
                "type": "array", "items": {"type": "string"},
                "description": "Tags for categorization.",
            },
        },
        "required": ["feedback_type", "content"],
    },
    fn=_t_save_feedback,
)

TOOL_REGISTRY["search_experience"] = ToolSpec(
    name="search_experience",
    description=(
        "Search past validation experiences and insights stored in the "
        "knowledge graph. Returns insights the agent has accumulated "
        "from previous sessions about specific fields or regulations."
    ),
    parameters={
        "type": "object",
        "properties": {
            "query": {"type": "string", "description": "Search query — field name, regulation, or topic."},
            "k": {"type": "integer", "default": 5},
        },
        "required": ["query"],
    },
    fn=_t_search_experience,
)

TOOL_REGISTRY["merge_nodes"] = ToolSpec(
    name="merge_nodes",
    description=(
        "Merge a duplicate insight node into a canonical one during sleep "
        "organization. Rewires edges and deletes the source node."
    ),
    parameters={
        "type": "object",
        "properties": {
            "source": {"type": "string", "description": "Duplicate node id to remove."},
            "target": {"type": "string", "description": "Canonical node id to keep."},
        },
        "required": ["source", "target"],
    },
    fn=_t_merge_nodes,
)

TOOL_REGISTRY["add_tags"] = ToolSpec(
    name="add_tags",
    description="Add categorization tags to an insight node.",
    parameters={
        "type": "object",
        "properties": {
            "node": {"type": "string", "description": "Insight node id."},
            "tags": {"type": "array", "items": {"type": "string"}},
        },
        "required": ["node", "tags"],
    },
    fn=_t_add_tags,
)

TOOL_REGISTRY["link_nodes"] = ToolSpec(
    name="link_nodes",
    description="Create a relationship edge between two graph nodes.",
    parameters={
        "type": "object",
        "properties": {
            "source": {"type": "string"},
            "target": {"type": "string"},
            "relation": {
                "type": "string",
                "default": "RELATES_TO",
                "description": "Edge type, e.g. RELATES_TO, SUPERSEDES.",
            },
        },
        "required": ["source", "target"],
    },
    fn=_t_link_nodes,
)

TOOL_REGISTRY["flag_conflict"] = ToolSpec(
    name="flag_conflict",
    description="Flag two contradictory insight nodes for human review.",
    parameters={
        "type": "object",
        "properties": {
            "node_a": {"type": "string"},
            "node_b": {"type": "string"},
            "reason": {"type": "string", "default": ""},
        },
        "required": ["node_a", "node_b"],
    },
    fn=_t_flag_conflict,
)

TOOL_REGISTRY["delete_node"] = ToolSpec(
    name="delete_node",
    description="Remove an orphan or irrelevant node from the knowledge graph.",
    parameters={
        "type": "object",
        "properties": {
            "node": {"type": "string", "description": "Node id to delete."},
        },
        "required": ["node"],
    },
    fn=_t_delete_node,
)

TOOL_REGISTRY["create_concept"] = ToolSpec(
    name="create_concept",
    description=(
        "Create a new Concept that abstracts multiple related insights into "
        "a stable semantic structure (EverMemOS-style consolidation). "
        "Use when ≥2 insights share a theme and lifting them reduces graph "
        "entropy. Links insights via CONTAINS edges."
    ),
    parameters={
        "type": "object",
        "properties": {
            "label": {"type": "string", "description": "Short concept name."},
            "definition": {"type": "string", "description": "Stable semantic summary."},
            "covers": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Insight node ids abstracted by this concept.",
            },
            "links_to": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Optional column/concept ids (e.g. LGD_ESTIMADA).",
            },
        },
        "required": ["label", "definition", "covers"],
    },
    fn=_t_create_concept,
)

TOOL_REGISTRY["link_to_concept"] = ToolSpec(
    name="link_to_concept",
    description="Attach an insight to an existing experience-derived Concept.",
    parameters={
        "type": "object",
        "properties": {
            "insight_id": {"type": "string"},
            "concept_id": {"type": "string"},
        },
        "required": ["insight_id", "concept_id"],
    },
    fn=_t_link_to_concept,
)

TOOL_REGISTRY["save_insight"] = ToolSpec(
    name="save_insight",
    description=(
        "Save a new insight or learning from the current validation session. "
        "The agent should call this when it discovers something useful — "
        "a quirk, an exception, or a pattern — so future sessions benefit."
    ),
    parameters={
        "type": "object",
        "properties": {
            "insight": {"type": "string", "description": "The insight text."},
            "related_fields": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Fields this insight relates to.",
            },
            "related_articles": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Regulation articles this insight relates to.",
            },
            "tags": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Tags for categorization.",
            },
        },
        "required": ["insight"],
    },
    fn=_t_save_insight,
)


def tool_schemas(names: list[str] | None = None) -> list[dict[str, Any]]:
    if names is None:
        return [t.schema() for t in TOOL_REGISTRY.values()]
    return [TOOL_REGISTRY[n].schema() for n in names if n in TOOL_REGISTRY]


def dispatch_tool(name: str, args: dict[str, Any]) -> dict[str, Any]:
    """Run a tool by name with the given JSON args; always returns a dict."""
    spec = TOOL_REGISTRY.get(name)
    if spec is None:
        return {"error": f"unknown_tool: {name}", "available": sorted(TOOL_REGISTRY)}
    # Sanitise: small models sometimes emit "null" as a string literal
    clean = {k: v for k, v in (args or {}).items() if v is not None and v != "null"}
    try:
        result = spec.fn(**clean)
    except TypeError as e:
        return {"error": f"bad_arguments: {e}", "tool": name, "args": args}
    except Exception as e:
        logger.exception("Tool %s failed", name)
        return {"error": f"runtime_error: {e!s}", "tool": name}
    if not isinstance(result, dict):
        result = {"result": result}
    return _truncate(result)
