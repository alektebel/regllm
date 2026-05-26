"""
Excel data dictionary extractor.

Reads .xlsx / .xls files where each row describes one database column.
Auto-detects header layout by matching known column name synonyms (ES + EN).
Returns (content_text, metadata) tuples ready for embedding and storage.
"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

_FIELD_SYNS    = {"campo", "column_name", "nombre_campo", "field", "variable", "nombre", "columna", "col"}
_DESC_SYNS     = {"descripcion", "descripción", "description", "definicion", "definición", "label", "etiqueta", "comentario", "comment"}
_TYPE_SYNS     = {"tipo", "type", "data_type", "tipo_dato", "formato", "format"}
_VALUES_SYNS   = {"valores", "valores_validos", "valores_válidos", "domain", "valid_values", "codes", "dominio", "catalogo", "catálogo", "posibles_valores"}
_EXAMPLE_SYNS  = {"ejemplo", "example", "sample", "ejemplo_valor", "muestra"}
_TABLE_SYNS    = {"tabla", "table", "table_name", "nombre_tabla", "entidad"}
_NULLABLE_SYNS = {"nullable", "nulo", "requerido", "required", "obligatorio"}


def _find_col(headers: list[str], synonyms: set[str]) -> str | None:
    for h in headers:
        if h.strip().lower() in synonyms:
            return h
    return None


def extract_excel(path: Path | str) -> list[tuple[str, dict]]:
    """Parse an Excel data dictionary.

    Returns a list of (content_text, metadata_dict) — one entry per column row.
    Both .xlsx and .xls are supported via openpyxl.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results: list[tuple[str, dict]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        field_col    = _find_col(headers, _FIELD_SYNS)
        desc_col     = _find_col(headers, _DESC_SYNS)
        type_col     = _find_col(headers, _TYPE_SYNS)
        values_col   = _find_col(headers, _VALUES_SYNS)
        example_col  = _find_col(headers, _EXAMPLE_SYNS)
        table_col    = _find_col(headers, _TABLE_SYNS)
        nullable_col = _find_col(headers, _NULLABLE_SYNS)

        if field_col is None and desc_col is None:
            logger.debug("Sheet %r: no recognisable field dict columns — skipping", sheet_name)
            continue

        for raw_row in rows[1:]:
            row = dict(zip(headers, raw_row))

            def cell(col: str | None) -> str:
                if col is None:
                    return ""
                v = row.get(col)
                return str(v).strip() if v is not None else ""

            field_name = cell(field_col)
            if not field_name:
                continue

            table_name   = cell(table_col) or sheet_name
            desc         = cell(desc_col)
            dtype        = cell(type_col)
            values       = cell(values_col)
            example      = cell(example_col)
            nullable_raw = cell(nullable_col)

            lines = [f"Tabla: {table_name} | Campo: {field_name}"]
            if dtype:
                type_line = f"Tipo: {dtype}"
                if nullable_raw:
                    required = nullable_raw.lower() not in ("yes", "sí", "si", "true", "1", "y")
                    type_line += f" | Requerido: {'Sí' if required else 'No'}"
                lines.append(type_line)
            if desc:
                lines.append(f"Descripción: {desc}")
            if values:
                lines.append(f"Valores válidos: {values}")
            if example:
                lines.append(f"Ejemplo: {example}")
            lines.append(f"Fuente: {path.name} (Hoja: {sheet_name})")

            metadata = {
                "source_type": "excel_dictionary",
                "table_name": table_name,
                "column_name": field_name.upper(),
                "source_file": path.name,
                "sheet_name": sheet_name,
                "confidence": "high",
            }
            results.append(("\n".join(lines), metadata))

    wb.close()
    logger.info("excel_dict: %d column entries from %s", len(results), path.name)
    return results
